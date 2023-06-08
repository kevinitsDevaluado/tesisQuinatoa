from django.contrib import messages
from django.contrib.auth.models import User
from django.shortcuts import render,redirect, HttpResponseRedirect
from django.core.urlresolvers import reverse_lazy
from apps.Libro.form import DocumentForm, DocumentFormDisabled
from django.views.generic import ListView,DeleteView, TemplateView
from apps.Libro.models import libro
from apps.campoAmplio.models import campoAmplio
from apps.campoEspecifico.models import campoEspecifico
from apps.autoresLibro.models import autoresLibro
from apps.baseDatos.models import baseDatos
from apps.palabraClave.models import palabraClave
from apps.roles.models import Rol
from apps.Investigador.models import Investigador
from apps.pais.models import pais
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import unicodedata

from openpyxl import Workbook
from django.http.response import HttpResponse

def elimina_tildes(cadena):
    s = ''.join((c for c in unicodedata.normalize('NFD', cadena) if unicodedata.category(c) != 'Mn'))
    return s

 #Creado por: Pilapaña, Guanotasig, Peralta, Pilatasig, Achote
 #Funcion para enviar correo al crear Nuevo Libro  Fecha: 23/07/2019
def correo(test1, test2):
    us = User.objects.get(id=test1)
    msg = MIMEMultipart()
    message = "El Sr(a). " + us.username + " a registrado un nuevo libro, con el Titulo:" + test2
    email_content = "El/La Investigador(a): " + us.username + " a registrado un nuevo Libro, con el Título: " + test2 + """
    <html>

<head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">

   <title>Ecuciencia</title>
   <style type="text/css">
    a {color: #943126;}
    b {color: #ffffff;font-size: 24px; font-family: Chercán}
  body, #header h1, #header h2, p {margin: 0; padding: 0;}
  #main {border: 1px solid #cfcece;}
  img {display: block;}
  #top-message p, #bottom p {color: #3f4042; font-size: 12px; font-family: Arial, Helvetica, sans-serif; }
  #header h1 {color: #f8f9f9 !important; font-family: "Lucida Grande", sans-serif; font-size: 24px; margin-bottom: 0!important; padding-bottom: 0; }
  #header p {color: ##F64C4C !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; font-size: 12px;  }
  h5 {margin: 0 0 0.8em 0;}
    h5 {font-size: 20px; color: ##F64C4C !important; font-family: Arial, Helvetica, sans-serif; }
  p {font-size: 12px; color: ##F64C4C !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; line-height: 1.5;}

   </style>
     <style type="text/css">
  .boton_personalizado{
    text-decoration: none;
    padding: 8px;
    font-weight: 600;
    font-size: 12px;
    color: #ffffff;
    background-color: #6388CD;
    border-radius: 6px;
  }
</style>
</head>

<body>


<table id="top-message" cellpadding="20" cellspacing="0" width="600" align="center">
    <tr>

    </tr>
  </table>

    <tr><table id="main" width="100" align="center" cellpadding="0" cellspacing="5" bgcolor="6388CD">

      <td align="center">

        <b>ECUCIENCIA</b>

        <table id="header"  cellpadding="10" cellspacing="0" align="center" bgcolor="F0F3F4">
          <tr>
             <td>
        <table>
          <tr>
            <td width="200" valign="top">
              <img src="http://ecuciencia.utc.edu.ec/static/images/logo4151.png" width="230" height="100">
            </td>
            <td width="15"></td>
            <td width="200" valign="top" align="center">
              <h5>Lista de Nuevos Libros</h5>
                  <a class="boton_personalizado" href="http://ecuciencia.utc.edu.ec/libro/listar">INGRESAR</a>
            </td>
          </tr>
        </table>
      </td>
          </tr>
        </table>
      </td>
    </tr>

  </table>
  <table id="button" cellpadding="20" cellspacing="0" width="600" align="center">
    <tr>
      <td align="center">
        <p>Este mensaje fue enviado a  por EcuCiencia</p>
        <p><a href="#">ecu.ciencia@utc.edu.ec</a> </p>
      </td>
    </tr>
  </table><!-- top message -->
</td></tr></table><!-- wrapper -->

</body>
    </html>


    """

    # setup the parameters of the message
    password = "3cuci3ncia.2018"
    msg['From'] = "ecu.ciencia@utc.edu.ec"
    msg['To'] = "alex.cevallos@utc.edu.ec, jessy.lema@utc.edu.ec"  # SECRETAIA DE INVESTIGACION 
    msg['Subject'] = "LIBRO: EL/LA INVESTIGADOR(A) " + us.first_name + " " + us.last_name + " REGISTRA UN NUEVO LIBRO"

    # add in the message body
    part1 = MIMEText(message, 'plain')
    part2 = MIMEText(email_content, 'html')

    msg.attach(part1)
    msg.set_payload(part2)

    # create server
    server = smtplib.SMTP('smtp.gmail.com: 587')

    server.starttls()

    # Login Credentials for sending the mail
    server.login(msg['From'], password)

    # send the message via the server.
    server.sendmail(msg['From'], msg['To'].split(","), msg.as_string())

    server.quit()

    print
    "successfully sent email to %s:" % (msg['To'])


class LibroList(ListView):
    model = libro
    template_name = 'libro/ListLibro.html'
    def get_context_data(self, **kwargs):
        context = super(LibroList, self).get_context_data(**kwargs)
        perfil = Investigador.objects.get(user_id=self.request.user.id)
        #book = libro.objects.all()
        us = User.objects.get(id=self.request.user.id)
        aut = autoresLibro.objects.filter(user=us)
        l = []
        for i in aut:
            l.append(i.libro.id)
        book = libro.objects.filter(id__in=l).order_by('Titulo')
        """
        var = []
        for art in book:
            for aut in art.autores.all():
                if aut.user == self.request.user:
                    # print(aut.user)
                    var.append(art.id)
        artUser = libro.objects.filter(id__in=var)
        """
        #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:03/07/2019
        context['bookAdmin'] = libro.objects.all().exclude(editableTrueFalse=1)
        context['bookAdminRec'] = libro.objects.all()
        context['perfil'] = perfil
        context['book'] = book
        return context

def Librocrear(request):
    usuario = request.user.id
    perfil = Investigador.objects.get(user_id=request.user.id)
    if request.method == 'POST':
        p = request.POST.getlist('palabras')

        userList=[]
        for i in User.objects.all():
            json = {}
            json['name'] =i.last_name + ' ' + i.first_name
            json['id'] = i.id
            userList.append(json)

        palabras = []
        for i in p:
            palabras.append(elimina_tildes(i.lower()))
        palabras = list(set(palabras))
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)
            book.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    book.PalabrasClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    book.PalabrasClave.add(palabra)

            for i in range(20):
                grado = "grado"+str(i)
                user = ''+str(i)
                if grado in request.POST:
                    usId = 0
                    for i in userList: 
                        if i['name'] == request.POST[user]:
                            usId = i['id'] 

                    if usId != 0:
                        a = request.POST[grado]
                        obj = User.objects.get(pk=int(usId))
                        if a != 'Primero':
                            autor = autoresLibro.objects.create(gradoAutoria=a, libro=book, user=obj, capituloSel=False)
                            autor.save()
                        else:
                            autor = autoresLibro.objects.create(gradoAutoria=a, libro=book, user=obj, capituloSel=True)
                            autor.save()
                        del autor
            book.user_id = usuario 
            titul = book.Titulo     
            book.save()

            book.LMtitulo = elimina_tildes(book.Titulo.upper())
            book.LMresumen = elimina_tildes(book.Resumen.upper())
            book.save()

            if request.POST['Aut'] == 'CoAutor':
                autorPos = autoresLibro.objects.get(libro_id = book.id, user_id = request.user.id)
                if "NumeroCap" or "Capitulo" in request.POST:
                    autorPos.capituloNumero = request.POST['NumeroCap']
                    autorPos.capituloTitulo = request.POST['Capitulo']
                    autorPos.capituloSel = False
                    autorPos.save()

            comprobacion = autoresLibro.objects.filter(libro_id = book.id)

            if comprobacion.count() == 0:
                autor = autoresLibro.objects.create(gradoAutoria="Primero", libro_id=book.id, user_id=book.user_id)
                autor.save()
                messages.warning(request, ('Se registro el libro, pero ocurrió un error con los autores revise la información nuevamente'))
                return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))

            correo(usuario,titul) #Llama a la funcion correo al guardar el libro
            messages.success(request, ('Se ha registrado el libro correctamente'))
            return redirect('Libro:libro_lis')
        else:
            messages.error(request, ('Ah ocurrido un error vuelva a intentarlo'))
            return redirect('libro/libro_crear')

    else:
        form = DocumentForm()
    return render(request,'libro/CreateLibro.html', {
        'form': form, 'base' : baseDatos.objects.all(), 
        'us': User.objects.exclude(is_superuser = True, is_staff = True).order_by('last_name'), 
        'perfil': perfil, 
        'amplio': campoAmplio.objects.all(), 
        'p': pais.objects.all().order_by('Nombre')
    })

def LibroEdit(request, idLibro):
    usuario = request.user.id
    perfil = Investigador.objects.get(user_id=usuario)
    libros = libro.objects.get(id=idLibro)

    users = [i.user_id for i in autoresLibro.objects.filter(libro=libros).order_by('id')]

    basesD = [val.id for val in libros.BaseDatos.all()]

    autores = autoresLibro.objects.filter(libro=libros)
    if not autoresLibro.objects.filter(libro=libros, user = usuario):
        permisos = 'No existe'
    else:
        gr1 = autoresLibro.objects.get(libro=libros, user=usuario)
        permisos = 'existe'

    if request.user.is_superuser:
        permisos = 'existe'
        libroAutor = autoresLibro.objects.filter(libro=libros).first()
    else:
            libroAutor = autoresLibro.objects.get(libro=libros, user_id=usuario)

    libObj = libro.objects.get(id=idLibro)
    palabra = [val.id for val in libObj.PalabrasClave.all()]
    p = request.POST.getlist('palabras')
    palabras = []
    for i in p:
        palabras.append(elimina_tildes(i.lower()))
    palabras = list(set(palabras))
    if request.method == 'POST':       

        userList = []
        for i in User.objects.all():
            json = {}
            json['name'] = i.last_name + ' ' + i.first_name
            json['id'] = i.id
            userList.append(json)

        if 'Capitulo' and 'NumeroCap' in request.POST:
            a = request.POST['Capitulo']
            b = request.POST['NumeroCap']
            libroAutor.capituloTitulo = request.POST['Capitulo']
            libroAutor.capituloNumero = request.POST['NumeroCap']
            libroAutor.save()
            print(libroAutor.capituloTitulo, libroAutor.id, libroAutor.user.id, a, b)
        form = DocumentForm(request.POST, request.FILES, instance=libros)
        if form.is_valid():

            book = form.save(commit=False)
            book.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    book.PalabrasClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    book.PalabrasClave.add(palabra)

            book.LMtitulo = elimina_tildes(book.Titulo.upper())
            book.LMresumen = elimina_tildes(book.Resumen.upper())
            book.statusText = False
            book.save()
            
            if book.textMining != None:
                book.textMining.statusClasificacion = False
                book.textMining.sublineaClasificacion = None
                book.textMining.save()
            
            au = autoresLibro.objects.filter(libro=book).order_by('id')
            for i in range(len(au)):
                grado = "grado"+str(i)
                user = ""+str(i)
                if grado in request.POST:
                    usId = 0
                    for j in userList:
                        if j['name'] == request.POST[user]:
                            usId = j['id']
                        if usId != 0:
                            au[i].gradoAutoria = request.POST[grado]
                            au[i].user = User.objects.get(pk=int(usId))
                            au[i].save()

            for i in range(len(au),20):
                grado = "grado"+str(i)
                user = ""+str(i)
                if grado in request.POST:
                    usId = 0
                    for i in userList:
                        if i['name'] == request.POST[user]:
                            usId = i['id']
                    if usId != 0:
                        a = request.POST[grado]
                        obj = User.objects.get(pk=int(usId))
                        autor = autoresLibro.objects.create(gradoAutoria=a, libro=book, user=obj, capituloSel=False)
                        autor.save()
                        del autor

            comprobacion = autoresLibro.objects.filter(libro_id = book.id)

            if comprobacion.count() == 0:
                autor = autoresLibro.objects.create(gradoAutoria="Primero", libro_id=book.id, user_id=book.user_id)
                autor.save()
                messages.warning(request, ('Se registro el libro, pero ocurrió un error con los autores revise la información nuevamente'))
                return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))

            messages.success(request, ('Se ha actualizado la información correctamente'))
            return redirect('Libro:libro_lis')
        else:
            messages.success(request, ('Ah ocurrido un error.'))
            return redirect('Libro:libro_lis')
    else:
        form = DocumentForm(instance=libros)
    if permisos == 'existe':
        return render(request, 'libro/UpdateLibro.html', {
            'form': form,
            'base': baseDatos.objects.all(),
            'Datos': baseDatos.objects.exclude(id__in=basesD),
            'palabra': palabraClave.objects.filter(id__in=palabra),
            'us': User.objects.exclude(is_superuser = True, is_staff = True).order_by('first_name'),
            'Autores': autoresLibro.objects.filter(libro=libros).order_by('id'),
            'users': User.objects.filter(is_superuser=False, is_staff=False, id__in=users),
            'id': idLibro,
            'perfil': perfil,
            'grad': libroAutor,
            'amplio': campoAmplio.objects.all(),
            'amId':libros.detallado.especifico.amplio.id,
            'esId':libros.detallado.especifico.id,
            'especifico': campoEspecifico.objects.all(),
            'p': pais.objects.all().order_by('Nombre')
        })
    else:
        return redirect('Libro:libro_lis')

def LibroEditDisabled(request, idLibro):
    usuario = request.user.id
    perfil = Investigador.objects.get(user_id=usuario)
    libros = libro.objects.get(id=idLibro)
    libroAutor = autoresLibro.objects.get(libro=libros, user_id=usuario)
    users = [i.user_id for i in autoresLibro.objects.filter(libro=libros).order_by('id')]
    print(libroAutor.id)

    basesD = [val.id for val in libros.BaseDatos.all()]

    autores = autoresLibro.objects.filter(libro=libros)
    if not autoresLibro.objects.filter(libro=libros, user = usuario):
        permisos = 'No existe'
    else:
        gr1 = autoresLibro.objects.get(libro=libros, user=usuario)
        permisos = 'existe'
    libObj = libro.objects.get(id=idLibro)
    palabra = [val.id for val in libObj.PalabrasClave.all()]
    p = request.POST.getlist('palabras')
    palabras = []
    for i in p:
        palabras.append(elimina_tildes(i.lower()))
    palabras = list(set(palabras))
    if request.method == 'POST':
        userList = []
        for i in User.objects.all():
            json = {}
            json['name'] = i.last_name + ' ' + i.first_name
            json['id'] = i.id
            userList.append(json)

        if 'Capitulo' and 'NumeroCap' in request.POST:
            a = request.POST['Capitulo']
            b = request.POST['NumeroCap']
            libroAutor.capituloTitulo = request.POST['Capitulo']
            libroAutor.capituloNumero = request.POST['NumeroCap']
            libroAutor.save()
            print(libroAutor.capituloTitulo, libroAutor.id, libroAutor.user.id, a, b)
        form = DocumentFormDisabled(request.POST, request.FILES, instance=libros)
        if form.is_valid():

            book = form.save(commit=False)
            book.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    book.PalabrasClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    book.PalabrasClave.add(palabra)

            book.LMtitulo = elimina_tildes(book.Titulo.upper())
            book.LMresumen = elimina_tildes(book.Resumen.upper())
            book.statusText = False
            book.save()
            
            if book.textMining != None:
                book.textMining.statusClasificacion = False
                book.textMining.sublineaClasificacion = None
                book.textMining.save()
            
            au = autoresLibro.objects.filter(libro=book).order_by('id')
            for i in range(len(au)):
                grado = "grado"+str(i)
                user = ""+str(i)
                if grado in request.POST:
                    usId = 0
                    for j in userList:
                        if j['name'] == request.POST[user]:
                            usId = j['id']
                        if usId != 0:
                            au[i].gradoAutoria = request.POST[grado]
                            au[i].user = User.objects.get(pk=int(usId))
                            au[i].save()

            for i in range(len(au),20):
                grado = "grado"+str(i)
                user = ""+str(i)
                if grado in request.POST:
                    usId = 0
                    for i in userList:
                        if i['name'] == request.POST[user]:
                            usId = i['id']
                    if usId != 0:
                        a = request.POST[grado]
                        obj = User.objects.get(pk=int(usId))
                        autor = autoresLibro.objects.create(gradoAutoria=a, libro=book, user=obj, capituloSel=False)
                        autor.save()
                        del autor

            comprobacion = autoresLibro.objects.filter(libro_id = book.id)

            if comprobacion.count() == 0:
                autor = autoresLibro.objects.create(gradoAutoria="Primero", libro_id=book.id, user_id=book.user_id)
                autor.save()
                messages.warning(request, ('Se registro el libro, pero ocurrió un error con los autores revise la información nuevamente'))
                return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))

            messages.success(request, ('Se ha actualizado la información correctamente'))
            return redirect('Libro:libro_lis')
        else:
            messages.success(request, ('Ah ocurrido un error.'))
            return redirect('Libro:libro_lis')
    else:
        form = DocumentFormDisabled(instance=libros)
    if permisos == 'existe':
        return render(request, 'libro/UpdateLibro.html', {
            'form': form,
            'base': baseDatos.objects.all(),
            'Datos': baseDatos.objects.exclude(id__in=basesD),
            'palabra': palabraClave.objects.filter(id__in=palabra),
            'us': User.objects.exclude(is_superuser = True, is_staff = True).order_by('first_name'),
            'Autores': autoresLibro.objects.filter(libro=libros).order_by('id'),
            'users': User.objects.filter(is_superuser=False, is_staff=False, id__in=users),
            'id': idLibro,
            'perfil': perfil,
            'grad': libroAutor,
            'amplio': campoAmplio.objects.all(),
            'amId':libros.detallado.especifico.amplio.id,
            'esId':libros.detallado.especifico.id,
            'especifico': campoEspecifico.objects.all(),
            'p': pais.objects.all().order_by('Nombre')
        })
    else:
        return redirect('Libro:libro_lis')

class LibroDelete(DeleteView):
    model = libro
    template_name = 'libro/DeleteLibro.html'
    success_url = reverse_lazy('Libro:libro_lis')
    def get_context_data(self, **kwargs):
        context = super(LibroDelete, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        perfil = Investigador.objects.get(user_id=usuario)
        roles = perfil.roles.all()
        privi = []
        privilegios = []
        privilegio = []
        for r in roles:
            privi.append(r.id)
        for p in privi:
            roles5 = Rol.objects.get(pk=p)
            priv = roles5.privilegios.all()
            for pr in priv:
                privilegios.append(pr.codename)
        for i in privilegios:
            if i not in privilegio:
                privilegio.append(i)

        if not  autoresLibro.objects.filter(libro_id = self.object.id, user_id = self.request.user.id):
            context['permisos'] = 'invalido'

        context['usuario'] = privilegio
        return context

class ReporteExcelLibro(TemplateView):
    def get(self,request,*args,**kwargs):
        librosL=libro.objects.all()
        wb=Workbook()
        ws=wb.active
        ws['B1']='REPORTE DE LIBROS'
        ws.merge_cells('B1:E1')
        ws['B3']='TÍTULO'
        ws['C3']='FECHA DE PUBLICACIÓN'
        ws['D3']='ESTADO'
        ws['E3']='ISBN'
        ws['F3']='EDITORIAL'
        ws['G3']='URL'
        cont=6
        for  arti in librosL:
            ws.cell(row=cont,column=2).value=arti.Titulo
            ws.cell(row=cont,column=3).value=arti.fechaPublicacion
            ws.cell(row=cont,column=4).value=arti.estado
            ws.cell(row=cont,column=5).value=arti.ISBN
            ws.cell(row=cont,column=6).value=arti.Editorial
            ws.cell(row=cont,column=7).value=arti.Url
            cont+=1

        nombre_archivo=" ReporteLibros.xlsx"
        response=HttpResponse(content_type="aplication/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

from apps.informacionLaboral.models import informacionLaboral 
from apps.Investigador.models import Investigador 
def consulta(request):

    lista = [i.user_id for i in Investigador.objects.filter(informacionLaboral__carrera_id = 1)]
    
    print('Investigadores carrera Informatica')
    print(Investigador.objects.filter(informacionLaboral__carrera_id = 1).query)
    
    print('Total autores')
    print(autoresLibro.objects.filter(user_id__in = lista).count())
    print(autoresLibro.objects.filter(user__in = lista).query)

    print('Total escritores de libros, grado autoria Primero')
    print(autoresLibro.objects.filter(user__in = lista, capituloSel = True).count())
    print(autoresLibro.objects.filter(user__in = lista, capituloSel = True).query)
    print('Total escritores de libros, grado autoria cualquiera, se suman los primeros')
    print(autoresLibro.objects.filter(user__in = lista, capituloTitulo = None).count())
    print(autoresLibro.objects.filter(user__in = lista, capituloTitulo = None).query)
    print('Total Escritores de libros por capitulo')
    print(autoresLibro.objects.filter(user__in = lista, capituloSel = False).exclude(capituloTitulo = None).count())
    print(autoresLibro.objects.filter(user__in = lista, capituloSel = False).exclude(capituloTitulo = None).query)

    #http://localhost/SEMINARIO2019
    #http://ecuciencia.utc.edu.ec/SEMINARIO2019/
    #print([i.user.email for i in informacionLaboral.objects.filter(carrera = 1)])

    #autoresLibro.objects.
    return HttpResponseRedirect('/index/')