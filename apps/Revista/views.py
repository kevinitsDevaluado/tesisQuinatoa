from django.contrib.auth.models import User
from django.shortcuts import render,redirect
from django.core.urlresolvers import reverse_lazy

from apps.Articulos_Cientificos.models import articulos_cientificos
from apps.Revista.form import DocumentForm
from apps.Revista.models import revista
from apps.baseDatos.models import baseDatos
from apps.tipoBaseDatos.models import tipoBaseDatos
from django.views.generic import ListView,DeleteView,TemplateView
from apps.Investigador.models import Investigador
from apps.roles.models import Rol
from django.http import JsonResponse
from django.http import HttpResponse
import json
from collections import Counter
""" ------------------------importar para el excel"""
from openpyxl import Workbook
from django.http.response import HttpResponse
# Create your views here.

def revCreate(request):
    Nombre = request.POST.get('Nombre')
    print('Mi base de datos', Nombre)

    if not revista.objects.filter(Nombre__iexact=Nombre):
        if request.method == 'POST':
            ISSN = request.POST.get('ISSN')
            base = request.POST.getlist('baseData[]')
            dataBase = []
            print(base)
            for i in base:
                dataBase.append(baseDatos.objects.get(id=i))
            for i in dataBase:
                print(i.BaseDatos)
            cuartil_revista  = request.POST.get('Cuartil_Pertenece')
            Factor_Impacto  = request.POST.get('Factor_Impacto')
            Url  = request.POST.get('Url')
            us = request.POST.get('user')
            validada = request.POST.get('validada')
            Observacion="None"

            USER = User.objects.get(id=int(us))
            rev = revista.objects.create(
                Nombre=Nombre,
                ISSN=ISSN,
                cuartil_revista = cuartil_revista,
                Factor_Impacto = Factor_Impacto,
                Url = Url,
                Observacion=Observacion,
                estado=validada,
            )
            rev.user = USER
            rev.save()
            for i in dataBase:
                rev.base.add(i)
            rev.save
            results = []
            revjson = {}
            revjson['text'] = rev.Nombre
            print(rev.Nombre)
            revjson['value'] = rev.id
            print(rev.id)
            results.append(revjson)
            data_json = json.dumps(results)
            mimetype = "application/json"
            return HttpResponse(data_json, mimetype)
        else:
            response = HttpResponse('Your message here', status=401)
            response['Content-Length'] = len(response.content)
            return response
    else:
        response = HttpResponse('Your message here', status=401)
        response['Content-Length'] = len(response.content)
        return response


def listSelectedItems(request):
    if request.method == 'POST':
        data = request.POST.getlist('datos[]')
        bd = revista.objects.all()
        l = []
        for i in bd:
            for j in i.base.all():
                #print('revista: ', i , ' Base de datos anexada: ',j.id, j.BaseDatos)
                for k in data:
                    #print('bases Sel ',k)
                    #print('b REV ',j.id)
                    if int(k) == int(j.id):
                        #print(i, i.id)
                        l.append(i.id)
        revistas = list(set(l))
        r = revista.objects.filter(id__in = revistas)
        results=[]
        doctor_json = {}
        doctor_json['text'] = '----------'
        doctor_json['value'] = ''
        doctor_json['estado']=''
        doctor_json['observacion']=''
        doctor_json['usuario']=''
        results.append(doctor_json)
        for i in r:
            doctor_json={}
            doctor_json['text']=i.Nombre
            doctor_json['value']=i.id
            doctor_json['estado']=i.estado
            doctor_json['observacion']=i.Observacion
            doctor_json['usuario']=i.user.first_name + i.user.last_name
            results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

############################################
def listRev(request):
    if request.is_ajax:
         #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:04/07/2019
        bd=revista.objects.all().exclude(estado=4)
        results=[]
        for i in bd:
            doctor_json={}
            doctor_json['text']=i.Nombre
            doctor_json['value']=i.id
            results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)


def listRevUp(request):
    if request.method == 'POST':
        # search=request.POST.get('start','')
        ArticuloID = request.POST.get('ArticuloID')
        articulo = articulos_cientificos.objects.get(id=ArticuloID)
        revD = [val.id for val in articulo.revista.all()]

        revSel = [val for val in articulo.revista.all()]

        rev = revista.objects.all()

        revNoSel = revista.objects.exclude(id__in=revD)

        results = []
        for i in revSel:
            doctor_json = {}
            doctor_json['text'] = i.Nombre
            doctor_json['value'] = i.id
            doctor_json['selected'] = 'selected'
            results.append(doctor_json)

        for i in revNoSel:
            doctor_json = {}
            doctor_json['text'] = i.Nombre
            doctor_json['value'] = i.id
            doctor_json['selected'] = ''
            results.append(doctor_json)

        data_json = json.dumps(results)

    else:
        data_json = 'fail'
    mimetype = "application/json"
    return HttpResponse(data_json, mimetype)

def Revistacrear(request):
    usuario = request.user.id
    perfil = Investigador.objects.get(user_id=usuario)
    roles = perfil.roles.all()
    privi = []
    privilegios = []
    privilegio = []
    for r in roles:
        privi.append(r.id)
    for p in privi:
        roles5 = Rol.objects.get(pk=p)
        priv = roles5.privilegios.all()
        for pr in priv:
            privilegios.append(pr.codename)
    for i in privilegios:
        if i not in privilegio:
            privilegio.append(i)
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('Revista:lista_Revista')
    else:
        form = DocumentForm()
    return render(request, 'revista/CreateRevista.html', {
        'form': form, 'usuario': privilegio
    })

class RevistaList(ListView):
    model = revista
    template_name = 'revista/ListRevista.html'
    def get_context_data(self, **kwargs):
        context = super(RevistaList, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        try:
            capa = revista.objects.filter(user_id=usuario)
        except revista.DoesNotExist:
            capa = None
        #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:03/07/2019
        context['checkl'] = 0
        context['Pon'] = capa
        context['revistaAdmin']=revista.objects.all().exclude(estado=4)
        context['revistaAdminRec']=revista.objects.all()
        context['perfil'] = Investigador.objects.get(user_id=self.request.user.id)
        return context
    

def RevistaEdit(request, id_revista):
    usuario = request.user.id
    articulo = revista.objects.get(id=id_revista)
    basesD = [val.id for val in articulo.base.all()]
    db1 = articulo.base.all().first()
    proy = revista.objects.get(id=id_revista)
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES,instance=proy)
        if form.is_valid():
            form.save()
            return redirect('Revista:lista_Revista')
    else:
        form = DocumentForm(instance=proy)
    return render(request, 'revista/UpdateRevista.html', {
        'DB1': db1.tipoBaseDatos_id,
        'perfil': Investigador.objects.get(user_id=request.user.id),
        'form': form,
        'Datos': baseDatos.objects.exclude(id__in=basesD),
        'base': baseDatos.objects.all()
    })

class RevistaDelete(DeleteView):
    model = revista
    template_name = 'revista/DeleteRevista.html'
    success_url = reverse_lazy('Revista:lista_Revista')
    def get_context_data(self, **kwargs):
        context = super(RevistaDelete, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        perfil = Investigador.objects.get(user_id=usuario)
        roles = perfil.roles.all()
        privi = []
        privilegios = []
        privilegio = []
        for r in roles:
            privi.append(r.id)
        for p in privi:
            roles5 = Rol.objects.get(pk=p)
            priv = roles5.privilegios.all()
            for pr in priv:
                privilegios.append(pr.codename)
        for i in privilegios:
            if i not in privilegio:
                privilegio.append(i)
        context['usuario'] = privilegio
        return context

"""----------ACTUALIZAR ESTADO DE LA REVISTA-------------------"""
def estadoA(request):    
    observacion=request.POST.get('Observacion')
    estado=request.POST.get('estado')
    id_revista=request.POST.get('id')
    model=revista
    print("hola mundi",estado,id_revista,observacion)
    proy = revista.objects.get(id=id_revista)
    if request.method == 'POST':
     revista.objects.filter(id=id_revista).update(
     estado=estado,
     Observacion=observacion,
     )
     print("ESTOY AQUI")
     r = revista.objects.all().filter(id = id_revista)
     results=[]
     doctor_json = {}
     doctor_json['text'] = ''
     doctor_json['validar']=''
     doctor_json['estado']=''
     results.append(doctor_json)
     for i in r:
        doctor_json={}
        doctor_json['text']=i.Nombre 
        doctor_json['validar']=i.validar
        doctor_json['estado']=i.estado
        
        results.append(doctor_json)
     data_json=json.dumps(results)    
    else:data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)
#-------------Para la notificación de nuemro de reistas rechazadas-----------#
def verRevista(request):
    usuario=request.GET.get('user')
    print(usuario)
    xd=usuario
    if usuario == '1':
        print("Admin")
    else:
        print("Normal")
    if request.method == 'GET':
        x=0
        if usuario == '1':
            r = revista.objects.all().filter(estado=4)   
        else:
            r = revista.objects.filter(user_id=usuario, estado=4)    
        results=[]
        for i in r:
            x=x+1
        doctor_json={}   
        doctor_json['estado']=x
        print(doctor_json['estado'])   
        results.append(doctor_json)
        data_json=json.dumps(results)    
    else:data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

class ReporteExcel(TemplateView):
    
    def get(self,request,*args,**kwargs):
        serparador=' '
        revistaLista=consultarRevistas()
        wb=Workbook()
        ws=wb.active
        ws['B1']='REPORTE DE REVISTAS'
        ws.merge_cells('B1:E1')
        ws['B3']='ID'
        ws['C3']='NOMBRE'
        ws['D3']='AUTOR'
        ws['E3']='URL'
        ws['F3']='CUARTIL'
        ws['G3']='FACTOR IMPACTO'
        ws['H3']='ESTADO'
        ws['I3']='TIPO DE REVISTA'
        ws['J3']='OBSERVACIÓN'
        cont=7
        for  rev in revistaLista:
            ws.cell(row=cont,column=2).value=rev.id
            ws.cell(row=cont,column=3).value=rev.Nombre
            ws.cell(row=cont,column=4).value=rev.first_name+" "+rev.last_name
            ws.cell(row=cont,column=5).value=rev.Url
            ws.cell(row=cont,column=6).value=rev.Cuartil_Pertenece
            #ws.cell(row=cont,column=7).value=rev.Factor_Impacto
            if rev.estado==1:
                ws.cell(row=cont,column=8).value="Ingresada"
            elif rev.estado==2:
                ws.cell(row=cont,column=8).value="Aceptada"
            elif rev.estado==3:
                ws.cell(row=cont,column=8).value="Corregida"
            elif rev.estado==4:
                ws.cell(row=cont,column=8).value="Rechazada"
            if rev.tipoBaseDatos.id==1:
                 ws.cell(row=cont,column=9).value="Regional"
            if rev.tipoBaseDatos.id==2:
                 ws.cell(row=cont,column=9).value="Alto Impacto"
            ws.cell(row=cont,column=10).value=rev.Observacion
            cont+=1

        nombre_archivo=" ReporteRevistas.xlsx"
        response=HttpResponse(content_type="aplication/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response
def consultarRevistas():
    query ='SELECT "Revista_revista".id, "Revista_revista"."Nombre",   "Revista_revista"."ISSN",   "Revista_revista"."Cuartil_Pertenece",   "Revista_revista"."Factor_Impacto", '
    query2= '  "Revista_revista"."Url",   "Revista_revista".estado,   "Revista_revista"."Observacion",   "baseDatos_basedatos"."tipoBaseDatos_id",    '
    query3= '   auth_user.first_name, auth_user.last_name,  COUNT(*) AS RecuentoFilas'
    query4= ' FROM   public."Revista_revista",   public."Revista_revista_base",   public."baseDatos_basedatos",   public."tipoBaseDatos_tipobasedatos",   public.auth_user'
    query5= ' WHERE "Revista_revista".id = "Revista_revista_base".revista_id AND'
    query6= ' "baseDatos_basedatos".id = "Revista_revista_base".basedatos_id AND'
    query7= ' "baseDatos_basedatos"."tipoBaseDatos_id" = "tipoBaseDatos_tipobasedatos".id AND'
    query8= ' auth_user.id = "Revista_revista".user_id'
    query9=' GROUP BY "Revista_revista".id,"Revista_revista"."Nombre","baseDatos_basedatos"."tipoBaseDatos_id", "Revista_revista"."ISSN", '
    query10='   "Revista_revista"."Cuartil_Pertenece",  "Revista_revista"."Factor_Impacto",   "Revista_revista"."Url",   "Revista_revista".estado,   "Revista_revista"."Observacion", '
    query11='   "baseDatos_basedatos"."tipoBaseDatos_id",     auth_user.first_name,   auth_user.last_name'
    query12=' HAVING COUNT(*) >= 0'
    query13=' ORDER BY public."Revista_revista".id'
    queryTotal=query+query2+query3+query4+query5+query6+query7+query8+query9+query10+query11+query12+query13
    #print(queryTotal)
    ar=baseDatos.objects.raw(queryTotal)    
    return(ar)