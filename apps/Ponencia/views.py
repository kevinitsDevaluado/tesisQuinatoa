from django.contrib.auth.models import User
from django.shortcuts import render
from django.core.urlresolvers import reverse_lazy
from django.contrib import messages
from django.shortcuts import render, HttpResponseRedirect
from apps.Ponencia.form import PonenciaForm,PonenciaFormDisabled
from apps.Ponencia.models import ponencia
from django.views.generic import ListView, CreateView,UpdateView,DeleteView,TemplateView
from apps.Investigador.models import Investigador
from apps.autoresPonencia.models import autoresPonencia
from apps.autoresArticulos.models import autoresArticulos
from apps.palabraClave.models import palabraClave
from apps.roles.models import Rol
from apps.Articulos_Cientificos.models import articulos_cientificos
from apps.pais.models import pais
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
# Create your views here.
import unicodedata

from apps.universidad.models import universidad
from openpyxl import Workbook
from django.http.response import HttpResponse

def elimina_tildes(cadena):
    s = ''.join((c for c in unicodedata.normalize('NFD', cadena) if unicodedata.category(c) != 'Mn'))
    return s

#Creado por: Pilapaña, Guanotasig, Peralta, Pilatasig, Achote
#Funcion para enviar correo al crear Nueva Ponencia Fecha: 23/07/2019
def correo(test1, test2):
    us = User.objects.get(id=test1)
    msg = MIMEMultipart()
    message = "El Sr(a). " + us.username  + " a registrado una nueva ponencia, con el Titulo:" + test2
    email_content = "El/La Investigador(a): " + us.username  + " a registrado una nueva Ponencia, con el Título: " + test2 + """
    <html>
  
 <head>
<meta http-equiv="Content-Type" content="text/html; charset=utf-8">

   <title>Ecuciencia</title>
   <style type="text/css">
    a {color: #943126;}
    b {color: #ffffff;font-size: 24px; font-family: Chercán}
  body, #header h1, #header h2, p {margin: 0; padding: 0;}
  #main {border: 1px solid #cfcece;}
  img {display: block;}
  #top-message p, #bottom p {color: #3f4042; font-size: 12px; font-family: Arial, Helvetica, sans-serif; }
  #header h1 {color: #f8f9f9 !important; font-family: "Lucida Grande", sans-serif; font-size: 24px; margin-bottom: 0!important; padding-bottom: 0; }
  #header p {color: ##F64C4C !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; font-size: 12px;  }
  h5 {margin: 0 0 0.8em 0;}
    h5 {font-size: 20px; color: ##F64C4C !important; font-family: Arial, Helvetica, sans-serif; }
  p {font-size: 12px; color: ##F64C4C !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; line-height: 1.5;}

   </style>
   <style type="text/css">
  .boton_personalizado{
    text-decoration: none;
    padding: 8px;
    font-weight: 600;
    font-size: 12px;
    color: #ffffff;
    background-color: #6388CD;
    border-radius: 6px;
  }
</style>
</head>

<body>


<table id="top-message" cellpadding="20" cellspacing="0" width="600" align="center">
    <tr>

    </tr>
  </table>

    <tr><table id="main" width="100" align="center" cellpadding="0" cellspacing="5" bgcolor="6388CD">

      <td align="center">

        <b>ECUCIENCIA</b>

        <table id="header"  cellpadding="10" cellspacing="0" align="center" bgcolor="F0F3F4">
          <tr>
             <td>
        <table>
          <tr>
            <td width="200" valign="top">
              <img src="http://ecuciencia.utc.edu.ec/static/images/logo4151.png" width="230" height="100">
            </td>
            <td width="15"></td>
            <td width="200" valign="top" align="center">
              <h5>Lista de Nuevas Ponencias </h5>
              <a class="boton_personalizado" href="http://ecuciencia.utc.edu.ec/ponencia/listar">INGRESAR </a>
            </td>
          </tr>
        </table>
      </td>
          </tr>
        </table>
      </td>
    </tr>

  </table>
  <table id="button" cellpadding="20" cellspacing="0" width="600" align="center">
    <tr>
      <td align="center">
        <p>Este mensaje fue enviado a  por EcuCiencia</p>
        <p><a href="#">ecu.ciencia@utc.edu.ec</a> </p>
      </td>
    </tr>
  </table><!-- top message -->
</td></tr></table><!-- wrapper -->

</body>
    </html>


    """

    # setup the parameters of the message
    password = "3cuci3ncia.2018"
    msg['From'] = "ecu.ciencia@utc.edu.ec"
    msg['To'] = "alex.cevallos@utc.edu.ec, jessy.lema@utc.edu.ec"  # SECRETAIA DE INVESTIGACION
    msg['Subject'] = "PONENCIA: EL/LA INVESTIGADOR(A) " + us.first_name + " " + us.last_name + " REGISTRA UNA NUEVA PONENCIA"

    # add in the message body
    part1 = MIMEText(message, 'plain')
    part2 = MIMEText(email_content, 'html')

    msg.attach(part1)
    msg.set_payload(part2)

    # create server
    server = smtplib.SMTP('smtp.gmail.com: 587')

    server.starttls()

    # Login Credentials for sending the mail
    server.login(msg['From'], password)

    # send the message via the server.
    server.sendmail(msg['From'], msg['To'].split(","), msg.as_string())

    server.quit()

    print
    "successfully sent email to %s:" % (msg['To'])


class PonenciaList(ListView):
    model = ponencia
    template_name = 'ponencia/ponencia_listar.html'
    def get_context_data(self, **kwargs):
        context = super(PonenciaList, self).get_context_data(**kwargs)
        us = User.objects.get(id=self.request.user.id)
        aut = autoresPonencia.objects.filter(user=us)
        l = []
        for i in aut:
            l.append(i.ponencia.id)
        pon = ponencia.objects.filter(id__in=l)
        context['perfil'] = Investigador.objects.get(user_id=self.request.user.id)
        context['Pon'] = pon
        #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:03/07/2019
        context['ponenciaAdmin'] = ponencia.objects.all().exclude(editableTrueFalse=1)
        context['ponenciaAdminRec'] = ponencia.objects.all()
        return context

class PonenciaCreate(CreateView):
    model = ponencia
    form_class = PonenciaForm
    template_name = 'ponencia/ponencia_crear.html'
    success_url = reverse_lazy('Ponencia:ponencia_listar')
    def get_context_data(self, **kwargs):
        context = super(PonenciaCreate, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        autPon = autoresArticulos.objects.filter(user_id = usuario)
        lPon = []
        for i in autPon:
            lPon.append(i.articulo.id)
        context['arti'] = articulos_cientificos.objects.filter(id__in = lPon).order_by('titulo')
        context['uni'] = universidad.objects.all()
        context['us'] = User.objects.exclude(is_superuser = True, is_staff = True).order_by('first_name')
        context['articulos'] = articulos_cientificos.objects.all()
        context['perfil'] = Investigador.objects.get(user_id=self.request.user.id)
        context['p'] = pais.objects.all().order_by('Nombre')
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        userList=[]
        for i in User.objects.all():
            json = {}
            json['name'] =i.last_name + ' ' + i.first_name
            json['id'] = i.id
            userList.append(json)
        
        usuario = self.request.user.id
        form = self.form_class(request.POST, request.FILES)
        p = request.POST.getlist('palabras')
        palabras = []
        for i in p:
            palabras.append(elimina_tildes(i.lower()))
        palabras = list(set(palabras))
        if form.is_valid():
            pon = form.save(commit=False)
            pon.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    pon.palabrasClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    pon.palabrasClave.add(palabra)
            #--------------------------------------------
            if "3" == request.POST["tipo"]:
                artPon = autoresArticulos.objects.filter(articulo = pon.articuloCientifico).order_by('id')
                for i in artPon:
                    autor = autoresPonencia.objects.create(gradoAutoria=i.gradoAutoria, ponencia=pon, user=i.user)
                    autor.save()
                    del autor
            else:
                for i in range(6):
                    grado = "grado"+str(i)
                    user = ""+str(i)
                    if grado in request.POST:
                        usId = 0
                        for i in userList:
                            if i['name'] == request.POST[user]:
                                usId = i['id']
                        if usId != 0:
                            a = request.POST[grado]
                            obj = User.objects.get(pk=int(usId))
                            autor = autoresPonencia.objects.create(gradoAutoria=a, ponencia=pon, user=obj)
                            autor.save()
                            del autor
            pon.user_id = usuario
            titul = pon.tituloPonencia
            pon.save()

            pon.PMnombre = elimina_tildes(pon.nombrePonencia.upper())
            pon.PMtitulo = elimina_tildes(pon.tituloPonencia.upper())
            pon.PMresumen = elimina_tildes(pon.resumen.upper())
            pon.save()
            correo(usuario,titul) #Llama a la funcion correo al guardar la ponencia

            messages.success(request, ('Se ha registrado la ponencia correctamente'))
            return HttpResponseRedirect(reverse_lazy('Ponencia:ponencia_listar'))
        else:
            messages.error(request, ('Por favor registrese nuevamente'))
            return self.render_to_response(self.get_context_data(form=form))


class PonenciaUpdate(UpdateView):

    model = ponencia
    form_class = PonenciaForm
    template_name = 'ponencia/ponencia_update.html'
    success_url = reverse_lazy('Ponencia:ponencia_listar')
    def get_context_data(self, **kwargs):
        context = super(PonenciaUpdate, self).get_context_data(**kwargs)

        ponenciaId = self.object.id
        
        usPon = User.objects.get(id = self.request.user.id)
        autPon = autoresArticulos.objects.filter(user = usPon)
        lPon = []
        for i in autPon:
            lPon.append(i.articulo.id)
        
        pon = ponencia.objects.get(id=ponenciaId)
        palabra = [val.id for val in pon.palabrasClave.all()]
        autores = autoresPonencia.objects.filter(ponencia=pon)
        users = [i.user_id for i in autoresPonencia.objects.filter(ponencia=pon).order_by('id')]
        
        context['articulos'] = articulos_cientificos.objects.filter(id__in = lPon).order_by('titulo')
        context['us'] = User.objects.exclude(is_superuser = True, is_staff = True).order_by('first_name')
        context['Autores'] = autoresPonencia.objects.filter(ponencia=pon).order_by('id')
        context['users'] = User.objects.filter(is_superuser=False, is_staff=False, id__in=users)
        context['palabra'] = palabraClave.objects.filter(id__in=palabra)
        context['uni'] = universidad.objects.all()
        context['p'] = pais.objects.all().order_by('Nombre')
        context['perfil'] = Investigador.objects.get(user_id=self.request.user.id)
        return context

    def post(self, request, *args, **kwargs):
        
        self.object = self.get_object()
        usuario = self.request.user.id
        ponId = self.object.id

        ponencia = self.model.objects.get(id=ponId)
        
        form = self.form_class(request.POST, request.FILES, instance=ponencia)
        p = request.POST.getlist('palabras')
        palabras = []
        userList = []
        for i in User.objects.all():
            json = {}
            json['name'] = i.last_name + ' ' + i.first_name
            json['id'] = i.id
            userList.append(json)
        for i in p:
            palabras.append(elimina_tildes(i.lower()))
        palabras = list(set(palabras))
        print(form)
        if form.is_valid():
            pon = form.save(commit=False)
            pon.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    pon.palabrasClave.add(palabraNew)
                    del palabraNew
                    print(i)
                else:
                    print(i)
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    pon.palabrasClave.add(palabra)
                    pon.save()
            pon.PMnombre = elimina_tildes(pon.nombrePonencia.upper())
            pon.PMtitulo = elimina_tildes(pon.tituloPonencia.upper())
            pon.PMresumen = elimina_tildes(pon.resumen.upper())
            pon.statusText = False
            pon.save()
            
            if pon.textMining != None:
                pon.textMining.statusClasificacion = False
                pon.textMining.sublineaClasificacion = None
                pon.textMining.save()
            
            au = autoresPonencia.objects.filter(ponencia=pon)
            for i in au:
                i.delete()
            if "3" == request.POST["tipo"]:
                artPon = autoresArticulos.objects.filter(articulo = pon.articuloCientifico).order_by('id')
                for i in artPon:
                    autor = autoresPonencia.objects.create(gradoAutoria=i.gradoAutoria, ponencia=pon, user=i.user)
                    autor.save()
                    del autor
            else:
                for i in range(6):
                    grado = "grado"+str(i)
                    user = ""+str(i)
                    if grado in request.POST:
                        usId = 0
                        for i in userList:
                            if i['name'] == request.POST[user]:
                                usId = i['id']
                        if usId != 0:
                            a = request.POST[grado]
                            obj = User.objects.get(pk=int(usId))
                            autor = autoresPonencia.objects.create(gradoAutoria=a, ponencia=pon, user=obj)
                            autor.save()
                            del autor

            messages.success(request, ('Se ha actualizado información correctamente'))
            return HttpResponseRedirect(reverse_lazy('Ponencia:ponencia_listar'))
        else:
            messages.error(request, ('Por favor revise la información del formulario'))
            return self.render_to_response(self.get_context_data(form = form))

class PonenciaUpdateDisabled(UpdateView):

    model = ponencia
    form_class = PonenciaFormDisabled
    template_name = 'ponencia/ponencia_update.html'
    success_url = reverse_lazy('Ponencia:ponencia_listar')
    def get_context_data(self, **kwargs):
        context = super(PonenciaUpdateDisabled, self).get_context_data(**kwargs)

        ponenciaId = self.object.id
        
        usPon = User.objects.get(id = self.request.user.id)
        autPon = autoresArticulos.objects.filter(user = usPon)
        lPon = []
        for i in autPon:
            lPon.append(i.articulo.id)
        
        pon = ponencia.objects.get(id=ponenciaId)
        palabra = [val.id for val in pon.palabrasClave.all()]
        autores = autoresPonencia.objects.filter(ponencia=pon)
        users = [i.user_id for i in autoresPonencia.objects.filter(ponencia=pon).order_by('id')]
        
        context['articulos'] = articulos_cientificos.objects.filter(id__in = lPon).order_by('titulo')
        context['us'] = User.objects.exclude(is_superuser = True, is_staff = True).order_by('first_name')
        context['Autores'] = autoresPonencia.objects.filter(ponencia=pon).order_by('id')
        context['users'] = User.objects.filter(is_superuser=False, is_staff=False, id__in=users)
        context['palabra'] = palabraClave.objects.filter(id__in=palabra)
        context['uni'] = universidad.objects.all()
        context['p'] = pais.objects.all().order_by('Nombre')
        context['perfil'] = Investigador.objects.get(user_id=self.request.user.id)
        return context

    def post(self, request, *args, **kwargs):
        
        self.object = self.get_object()
        usuario = self.request.user.id
        ponId = self.object.id

        ponencia = self.model.objects.get(id=ponId)
        
        form = self.form_class(request.POST, request.FILES, instance=ponencia)
        p = request.POST.getlist('palabras')
        palabras = []
        userList = []
        for i in User.objects.all():
            json = {}
            json['name'] = i.last_name + ' ' + i.first_name
            json['id'] = i.id
            userList.append(json)
        for i in p:
            palabras.append(elimina_tildes(i.lower()))
        palabras = list(set(palabras))
        print(form)
        if form.is_valid():
            pon = form.save(commit=False)
            pon.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    pon.palabrasClave.add(palabraNew)
                    del palabraNew
                    print(i)
                else:
                    print(i)
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    pon.palabrasClave.add(palabra)
                    pon.save()
            pon.PMnombre = elimina_tildes(pon.nombrePonencia.upper())
            pon.PMtitulo = elimina_tildes(pon.tituloPonencia.upper())
            pon.PMresumen = elimina_tildes(pon.resumen.upper())
            pon.statusText = False
            pon.save()
            
            if pon.textMining != None:
                pon.textMining.statusClasificacion = False
                pon.textMining.sublineaClasificacion = None
                pon.textMining.save()
                
            au = autoresPonencia.objects.filter(ponencia=pon)
            for i in au:
                i.delete()
            if "3" == request.POST["tipo"]:
                artPon = autoresArticulos.objects.filter(articulo = pon.articuloCientifico).order_by('id')
                for i in artPon:
                    autor = autoresPonencia.objects.create(gradoAutoria=i.gradoAutoria, ponencia=pon, user=i.user)
                    autor.save()
                    del autor
            else:
                for i in range(6):
                    grado = "grado"+str(i)
                    user = ""+str(i)
                    if grado in request.POST:
                        usId = 0
                        for i in userList:
                            if i['name'] == request.POST[user]:
                                usId = i['id']
                        if usId != 0:
                            a = request.POST[grado]
                            obj = User.objects.get(pk=int(usId))
                            autor = autoresPonencia.objects.create(gradoAutoria=a, ponencia=pon, user=obj)
                            autor.save()
                            del autor

            messages.success(request, ('Se ha actualizado información correctamente'))
            return HttpResponseRedirect(reverse_lazy('Ponencia:ponencia_listar'))
        else:
            messages.error(request, ('Por favor revise la información del formulario'))
            return self.render_to_response(self.get_context_data(form = form))

class PonenciaDelete(DeleteView):
    model = ponencia
    template_name = 'ponencia/ponencia_eliminar.html'
    success_url = reverse_lazy('Ponencia:ponencia_listar')
    def get_context_data(self, **kwargs):
        context = super(PonenciaDelete, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        perfil = Investigador.objects.get(user_id=usuario)
        roles = perfil.roles.all()
        privi = []
        privilegios = []
        privilegio= []
        for r in roles:
            privi.append(r.id)
        for p in privi:
            roles5 = Rol.objects.get(pk=p)
            priv = roles5.privilegios.all()
            for pr in priv:
                privilegios.append(pr.codename)
        for i in privilegios:
            if i not in privilegio:
                privilegio.append(i)
        context['usuario'] = privilegio
        return context

class ReporteExcelPonencia(TemplateView):
    def get(self,request,*args,**kwargs):
        ponenciaL=ponencia.objects.all()
        wb=Workbook()
        ws=wb.active
        ws['B1']='REPORTE DE PONENCIAS'
        ws.merge_cells('B1:E1')
        ws['B3']='NOMBRE'
        ws['C3']='LUGAR'
        ws['D3']='TÍTULO'
        ws['E3']='FECHA'
        ws['F3']='ISBN'
        ws['G3']='URL'
        cont=6
        for  arti in ponenciaL:
            ws.cell(row=cont,column=2).value=arti.nombrePonencia
            ws.cell(row=cont,column=3).value=arti.lugarPonencia
            ws.cell(row=cont,column=4).value=arti.tituloPonencia
            ws.cell(row=cont,column=5).value=arti.fechaPonencia
            ws.cell(row=cont,column=6).value=arti.isbn
            ws.cell(row=cont,column=7).value=arti.urlPonencia
            cont+=1

        nombre_archivo=" ReportePonencias.xlsx"
        response=HttpResponse(content_type="aplication/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response
