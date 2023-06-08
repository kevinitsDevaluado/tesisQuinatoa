from django.contrib.auth.models import User
from django.http import JsonResponse
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseForbidden
from django.shortcuts import render, render_to_response
from django.template import RequestContext

from apps.Articulos_Cientificos.models import articulos_cientificos
from apps.universidad.models import universidad
from apps.baseDatos.models import baseDatos
from apps.pais.models import pais
"""Imporart """
from django.views.generic import ListView,TemplateView
from apps.Investigador.models import Investigador
from apps.baseDatos.form import baseForm
import json

""" ------------------------importar para el excel"""
from openpyxl import Workbook
from django.http.response import HttpResponse


def listBD(request):
    if request.is_ajax:
        #search=request.POST.get('start','')
        #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:04/07/2019
        bd=baseDatos.objects.filter(tipoBaseDatos_id = 1).exclude(estado=3)
        results=[]
        for i in bd:
            doctor_json={}
            doctor_json['text']=i.BaseDatos
            doctor_json['value']=i.id
            results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

def nuevaDB(request):
    print('hola')
    if request.is_ajax:
         #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:04/07/2019
        bd=baseDatos.objects.filter(tipoBaseDatos_id = 2).exclude(estado=3)
        results=[]
        for i in bd:
            doctor_json={}
            doctor_json['text']=i.BaseDatos
            doctor_json['value']=i.id
            results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

def existe(request):
    bd=pais.objects.all()
    results=[]
    for i in bd:
        doctor_json={}
        doctor_json['nombre']=i.Nombre
        doctor_json['iso']=i.Iso
        results.append(doctor_json)
    data_json=json.dumps(results)
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)


def listDbUp(request):
    if request.method == 'POST':
        # search=request.POST.get('start','')
        ArticuloID = request.POST.get('ArticuloID')
        articulo = articulos_cientificos.objects.get(id=ArticuloID)

        basesD = [val.id for val in articulo.baseDatos.all()]

        dbSel = [val for val in articulo.baseDatos.all()]

        bd = baseDatos.objects.all()

        dbNoSel = baseDatos.objects.exclude(id__in=basesD)

        results = []
        for i in dbSel:
            doctor_json = {}
            doctor_json['text'] = i.BaseDatos
            doctor_json['value'] = i.id
            doctor_json['selected'] = 'selected'
            results.append(doctor_json)

        for i in dbNoSel:
            doctor_json = {}
            doctor_json['text'] = i.BaseDatos
            doctor_json['value'] = i.id
            doctor_json['selected'] = ''
            results.append(doctor_json)

        data_json = json.dumps(results)

    else:
        data_json = 'fail'
    mimetype = "application/json"
    return HttpResponse(data_json, mimetype)

def createBD(request):
    results = []
    dbjson = {}
    BaseDatos = request.POST.get('BaseDatos')

    print('Mi base de datos', BaseDatos)
    if not baseDatos.objects.filter(BaseDatos__iexact = BaseDatos):
        if request.method == 'POST':
            Url = request.POST.get('Url')
            tipoBD = request.POST.get('tipoBD')
            us = request.POST.get('user')
            USER = User.objects.get(id=int(us))
            newDB = baseDatos.objects.create(
                BaseDatos=BaseDatos,
                Url=Url,
                user= USER,
                tipoBaseDatos_id = tipoBD,
                validar = 'registrada',
                estado=1,
            )
            dbjson['text'] = newDB.BaseDatos
            dbjson['value'] = newDB.id
            results.append(dbjson)
            data_json = json.dumps(results)
            mimetype = "application/json"
            return HttpResponse(data_json, mimetype)
        else:
            response = HttpResponse('Your message here', status=401)
            response['Content-Length'] = len(response.content)
            return response
    else:
        response = HttpResponse('Your message here', status=401)
        response['Content-Length'] = len(response.content)
        return response


def listBDUniversidad(request):
    print('Instituto')
    dbjson = {}
    if request.is_ajax:
        p = request.POST.get('pais')
        uni = universidad.objects.filter(pais_id=p, instituto = 'SI').order_by('Nombre')
        uniOtras = universidad.objects.get(id = 1852)
        results=[]
        doctor_json = {}
        doctor_json['text'] = '-----------'
        doctor_json['value'] = ''
        results.append(doctor_json)
        for i in uni:
            doctor_json={}
            doctor_json['text']=i.Nombre
            doctor_json['value']=i.id
            results.append(doctor_json)
        doctor_json = {}
        doctor_json['text'] = uniOtras.Nombre
        doctor_json['value'] = uniOtras.id
        results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

def listBDInstituto(request):
    print('Universidad')
    if request.is_ajax:
        p = request.POST.get('pais')
        uni = universidad.objects.filter(pais_id=p, instituto = 'NO').order_by('Nombre')
        uniOtras = universidad.objects.get(id = 1852)
        results=[]
        doctor_json = {}
        doctor_json['text'] = '-----------'
        doctor_json['value'] = ''
        results.append(doctor_json)
        for i in uni:
            doctor_json={}
            doctor_json['text']=i.Nombre
            doctor_json['value']=i.id
            results.append(doctor_json)
        doctor_json = {}
        doctor_json['text'] = uniOtras.Nombre
        doctor_json['value'] = uniOtras.id
        results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

def listUni(request):
    print('Universidad')
    if request.is_ajax:
        p = request.POST['pais']
        print(p)
        uni = universidad.objects.filter(pais_id=p, instituto = 'NO').order_by('Nombre')
        results=[]
        doctor_json = {}
        doctor_json['text'] = '-----------'
        doctor_json['value'] = ''
        results.append(doctor_json)
        for i in uni:
            doctor_json={}
            doctor_json['text']=i.Nombre
            doctor_json['value']=i.id
            results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

def listUniAll(request):
    if request.is_ajax:
        p = request.POST['pais']
        print(p)
        uni = universidad.objects.filter(pais_id=p).order_by('Nombre')
        results=[]
        doctor_json = {}
        doctor_json['text'] = '-----------'
        doctor_json['value'] = ''
        results.append(doctor_json)
        for i in uni:
            doctor_json={}
            doctor_json['text']=i.Nombre
            doctor_json['value']=i.id
            results.append(doctor_json)
        data_json=json.dumps(results)
    else:
        data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

"""------------------ LISTAR BASE DE DATOS"""
class BddList(ListView):
    model = baseDatos
    template_name = 'BaseDeDatos/ListBDD.html'
    def get_context_data(self, **kwargs):
        context = super(BddList, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        try:
            capa = baseDatos.objects.filter(user_id=usuario)
        except baseDatos.DoesNotExist:
            capa = None
        #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:04/07/2019
        context['checkl'] = 0
        context['Pon'] = capa
        context['bddAdmin'] = baseDatos.objects.all().exclude(estado=3)
        context['bddAdminRec'] = baseDatos.objects.all()
        context['perfil'] = Investigador.objects.get(user_id=self.request.user.id)
        return context

def estadoBdd(request):  
     
    estado=request.POST.get('estado')
    id_baseDatos=request.POST.get('id')
    model=baseDatos
    print("BASE DE DATOS", estado,id_baseDatos) 
    proy = baseDatos.objects.get(id=id_baseDatos)
    if request.method == 'POST':
     baseDatos.objects.filter(id=id_baseDatos).update(
     estado=estado,
     )
     print("ESTOY AQUI")
     r = baseDatos.objects.all().filter(id = id_baseDatos)
     results=[]
     for i in r:
        doctor_json={}   
        doctor_json['estado']=i.estado
        a=doctor_json['estado']=i.estado
        print("bsbs")
        print(a)
        results.append(doctor_json)
     data_json=json.dumps(results)    
    else:data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

def actualizarBdd(request):    
    
    BaseDatos=request.POST.get('BaseDatos')
    Url=request.POST.get('Url')
    id_bdd=request.POST.get('id')
    tipoBdd=request.POST.get('tipoBaseDatos')
    model=baseDatos
    print("hola mundi",BaseDatos,id_bdd,Url,tipoBdd)
    proy = baseDatos.objects.get(id=id_bdd)
    if request.method == 'POST':
     baseDatos.objects.filter(id=id_bdd).update(
     BaseDatos=BaseDatos,
     Url=Url,
     tipoBaseDatos_id=tipoBdd,
     )
     print("ESTOY AQUI")
     r = baseDatos.objects.all().filter(id = id_bdd)
     results=[]

     for i in r:
        doctor_json={}
        doctor_json['BaseDatos']=i.BaseDatos 
        doctor_json['Url']=i.Url
        doctor_json['estado']=i.tipoBaseDatos_id
        a=doctor_json['text']=i.BaseDatos
        b=doctor_json['validar']=i.Url
        c=doctor_json['estado']=i.tipoBaseDatos_id
        print(a,b,c)
        results.append(doctor_json)
     data_json=json.dumps(results)    
    else:data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

class ReporteExcel(TemplateView):
    
    def get(self,request,*args,**kwargs):
        serparador=' '
        bddl=baseDatos.objects.all()
        wb=Workbook()
        ws=wb.active
        ws['B1']='REPORTE DE BASE DE DATOS'
        ws.merge_cells('B1:E1')
        ws['B3']='ID'
        ws['C3']='NOMBRE'
        ws['D3']='AUTOR'
        ws['E3']='URL'
        ws['F3']='ESTADO'
        ws['G3']='TIPO DE BASE DE DATOS'
        cont=7
        for  bdd in bddl:
            ws.cell(row=cont,column=2).value=bdd.id
            ws.cell(row=cont,column=3).value=bdd.BaseDatos
            ws.cell(row=cont,column=4).value=bdd.user.first_name+" "+bdd.user.last_name
            ws.cell(row=cont,column=5).value=bdd.Url
            if bdd.estado==1:
                ws.cell(row=cont,column=6).value="Ingresada"
            elif bdd.estado==2:
                ws.cell(row=cont,column=6).value="Aceptada"
            elif bdd.estado==3:
                ws.cell(row=cont,column=6).value="Corregida"
            elif bdd.estado==4:
                ws.cell(row=cont,column=6).value="Rechazada"
            ws.cell(row=cont,column=7).value=bdd.tipoBaseDatos.Nombre
            cont+=1

        nombre_archivo=" ReporteBaseDeDatos.xlsx"
        response=HttpResponse(content_type="aplication/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response