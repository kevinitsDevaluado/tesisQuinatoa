from django.contrib import messages
import unicodedata
from django.shortcuts import render, HttpResponseRedirect, redirect
from apps.Investigador.models import Investigador
from apps.Investigador.form import cambioForm, PasswordForm, document
from apps.roles.models import Rol
from django.contrib.auth.models import User
from apps.Formacion_Academica.models import formacion_academica
from apps.Formacion_Academica.form import FormacionAform
from apps.Linea_Investigacion.models import linea_investigacion
from apps.Articulos_Cientificos.models import articulos_cientificos
from apps.Libro.models import libro
from apps.Ponencia.models import ponencia
from apps.palabraClave.models import palabraClave
from apps.autoresArticulos.models import autoresArticulos
from apps.autoresArticulos.models import autoresArticulos
from django.http import HttpResponse
from apps.Revista.models import revista
from apps.autoresPonencia.models import autoresPonencia

#add esto tambien

from apps.SisRec.models import GlobalKeywordsInvestigador
from apps.SisRec.models import GlobalKeywords
from apps.SisRec.models import ProyectosMacro
from apps.SisRec.models import Publicacion



import json
#Librerias agregado por: Quinchimbla; Azogue; Caiza; Guilcazo; Potosi; Jaya
#Librerias para el botòn excel  fecha: 23/07/2019

#Activar cuando la base de datos de Capitulos este habilitada en ECUCIENCIA
"""from apps.Capitulos.models import Capitulos
from apps.autoresCapitulos.models import autoresCapitulos"""

from apps.pais.models import pais
from apps.ciudad.models import ciudad
from apps.autoresLibro.models import autoresLibro
from apps.campoAmplio.models import campoAmplio
from django.views.generic import TemplateView
from apps.autoresPonencia.models import autoresPonencia
from openpyxl import  Workbook,load_workbook
from django.http.response import HttpResponse
from openpyxl.styles import Alignment,Font,Border
from openpyxl.chart import title
import os
import collections

def elimina_tildes(cadena):
    s = ''.join((c for c in unicodedata.normalize('NFD', cadena) if unicodedata.category(c) != 'Mn'))
    return s

def inicio(request):
    usuario = request.user.id
   
    perfil = Investigador.objects.get(user_id=usuario)
    investigador_id=perfil.id
    form_Academico = FormacionAform
    try:
        Formacion = formacion_academica.objects.filter(user_id=usuario)
    except formacion_academica.DoesNotExist:
        Formacion = None
    usuario1 = User.objects.get(id = usuario)
    roles = perfil.roles.all()
    privi = []
    privilegios = []
    privilegio = []
    for r in roles:
        privi.append(r.id)
    for p in privi:
        roles5 = Rol.objects.get(pk=p)
        priv = roles5.privilegios.all()
        for pr in priv:
            privilegios.append(pr.codename)
    for i in privilegios:
        if i not in privilegio:
            privilegio.append(i)
    if request.method == 'GET':
        form = cambioForm(instance=perfil)
        form_second = PasswordForm(instance=usuario1)
        form_four = document(instance=perfil)
        #print("No Llegue :(")
    else:
        form = cambioForm(request.POST, instance=perfil)
        form_second = PasswordForm(request.POST, instance=usuario1)
        form_third = FormacionAform(request.POST, request.FILES)
        form_four = document(request.POST, request.FILES, instance=perfil)
        if 'val' in request.POST:

            variable = request.POST['val']
            print(variable)

            if int(variable) == 2:
                if form_third.is_valid():
                    print('aqui')
                    For_aca = form_third.save()
                    For_aca.save()

                    For_aca.FAMdescripcion = elimina_tildes(For_aca.descripcion.upper())
                    For_aca.FAMtitulo = elimina_tildes(For_aca.titulo.upper())
                    For_aca.FAMtipoTitulo = elimina_tildes(For_aca.tipoTitulo.upper())
                    For_aca.save()

                    messages.success(request, ('Se ha guardado la del título información correctamente'))
                    return redirect('inicio:logeo')
                else:
                    descripcion = request.POST['descripcion']
                    anio = request.POST['anio']
                    pais = request.POST['pais']
                    nombreCentroEstudios = int(request.POST['nombreCentroEstudios'])
                    titulo = request.POST['titulo']
                    tipoTitulo = request.POST['tipoTitulo']
                    user = request.POST['user']
                    print(descripcion)
                    FAMdescripcion = elimina_tildes(descripcion.upper())
                    FAMtitulo = elimina_tildes(titulo.upper())
                    FAMtipoTitulo = elimina_tildes(tipoTitulo.upper())
                    print(FAMdescripcion)

                    formacion_academica.objects.create(
                        descripcion = descripcion,
                        anio =int(anio),
                        pais_id = int(pais),
                        nombreCentroEstudios_id = int(nombreCentroEstudios),
                        titulo = titulo,
                        tipoTitulo = tipoTitulo,
                        user_id = int(user),
                        FAMdescripcion = FAMdescripcion,
                        FAMtitulo = FAMtitulo,
                        FAMtipoTitulo = FAMtipoTitulo,
                    )


                    messages.success(request, ('Se ha guardado la del título información correctamente'))
                    return redirect('inicio:logeo')
            if int(variable) == 1:
                if form_four.is_valid():
                    form_four.save()
                    messages.success(request, ('Se ha guardado la información del documento correctamente'))
                    return redirect('inicio:logeo')

        if 'password' in request.POST and 'password1' in request.POST:
            password1 = request.POST['password']
            password2 = request.POST['password1']
            if form.is_valid() and form_second.is_valid() and password1 == password2:
                perfil = form.save()
                perfil.user = form_second.save()
                perfil.user.set_password(perfil.user.password)
                perfil.save()
                form_second.save()
            else:
                messages.error(request, ('Las contraseñas no coinciden, vuelva a intentarlo'))
                return render(request, 'usuario/newPass.html', {'form': form, 'form2': form_second})

        return redirect('inicio:logeo')
    if not perfil.cambio:
        return render(request, 'usuario/newPass.html', {'form': form, 'form2': form_second})
    else:
        #agregado esto tambien

       
       
        if (GlobalKeywordsInvestigador.objects.filter(investigador_id= investigador_id,estado=True).exists()):
            
            keywords=GlobalKeywords.objects.all().order_by("termino")
        
            #proyectosMacro=ProyectosMacro.objects.filter()
            if Publicacion.objects.filter(investigador_id=investigador_id).exists():
                publicacion=Publicacion.objects.filter(investigador_id=investigador_id)
                return render(request, 'base1/inicio.html', {'publicacion':publicacion,'keywords':keywords,'usuario': privilegio, 'perfil': perfil,'Formacion':Formacion, 'formAca': form_Academico, 'formDoc':form_four, 'usAll': User.objects.filter(is_active = True).order_by("email"),})
            else:
                return render(request, 'base1/inicio.html', {'publicacion':[],'keywords':keywords,'usuario': privilegio, 'perfil': perfil,'Formacion':Formacion, 'formAca': form_Academico, 'formDoc':form_four, 'usAll': User.objects.filter(is_active = True).order_by("email"),})

        else:
            if Publicacion.objects.filter(investigador_id=investigador_id).exists():
                publicacion=Publicacion.objects.filter(investigador_id=investigador_id)
                return render(request, 'base1/inicio.html', {'publicacion':publicacion,'keywords':[],'usuario': privilegio, 'perfil': perfil,'Formacion':Formacion, 'formAca': form_Academico, 'formDoc':form_four, 'usAll': User.objects.filter(is_active = True).order_by("email"),})
            else:
                return render(request, 'base1/inicio.html', {'publicacion':[],'keywords':[],'usuario': privilegio, 'perfil': perfil,'Formacion':Formacion, 'formAca': form_Academico, 'formDoc':form_four, 'usAll': User.objects.filter(is_active = True).order_by("email"),})
"""
import os
import time
from datetime import datetime
import timestring
from django.utils import timezone
from dateutil import parser
"""
def new(request):
    autoresArticulos.object.create(user = 457, articulo=102, gradoAutoria="Primero")
    """
    art = ponencia.objects.all()
    #certificado
    for i in art:
        if str(i.informe) != "":

            url = 'C:/workspace/Cienciometrico/media/'+str(i.informe)
            (mode, ino, dev, nlink, uid, gid, size, atime, mtime, ctime) = os.stat(url)

            from_input = parser.parse(time.ctime(mtime))
            print(from_input)
            print(type(from_input))
            print('---------------------')
            i.uploaded_at = from_input
            i.save()
            #Mon Aug 27 03:44:56 2018
            '''
            print(time.ctime(mtime))
            print(type(time.ctime(mtime)))
            print(i.uploaded_at)
            print(type(i.uploaded_at))
            print(url)
            print(datetime_object)
            '''
    """
    '''
    art = articulos_cientificos.objects.all()
    for i in art:
        i.AMtitulo = elimina_tildes(i.titulo.upper())
        i.AMresumen = elimina_tildes(i.resumen.upper())
        i.save()
        print("Ready")

    lib = libro.objects.all()
    for i in lib:
        i.LMtitulo = elimina_tildes(i.Titulo.upper())
        i.LMresumen = elimina_tildes(i.Resumen.upper())
        i.save()
        print("Ready")

    pon = ponencia.objects.all()
    for i in pon:
        i.PMnombre = elimina_tildes(i.nombrePonencia.upper())
        i.PMtitulo = elimina_tildes(i.tituloPonencia.upper())
        i.PMresumen = elimina_tildes(i.resumen.upper())
        i.save()
        print("Ready")

    pal = palabraClave.objects.all()
    for i in pal:
        i.PCMtermino = elimina_tildes(i.Termino.upper())
        i.save()
    '''

    return HttpResponseRedirect('/index/')

def oldNew(request):

    #autoresArticulos.object.create(user_id = 457, articulo_id=102, gradoAutoria="Primero")

    return HttpResponseRedirect('/index/')

#-------------Para la notificación de nuemro de documentos rechazadas-----------#
def verRechazados(request):
    usuario=request.GET.get('user')
    print(usuario)
    xd=usuario
    if usuario == '1':
        print("Admin")
    else:
        print("Normal")
    if request.method == 'GET':
        x=0
        ar=0
        poc=0

        #funcion para Contar para las revistas rechazadas
        r = revista.objects.filter(user_id=usuario, estado=4).count()
        print("el numero de revistas rechazadas son: ",r)

        #Funciones para contar el numero de articulos rechazados
        autArticulos = autoresArticulos.objects.filter(user_id=usuario)
        for art in autArticulos:
            a=articulos_cientificos.objects.filter(pk=art.articulo_id, editableTrueFalse=1)
            for arti in a:
                ar=ar+1
        print("el numero de articulos rechazados son: ",ar)

        #Funciones para contar el numero de ponencias rechazadas
        autPonencias = autoresPonencia.objects.filter(user_id=usuario)
        for art in autPonencias:
            p=ponencia.objects.filter(pk=art.ponencia_id, editableTrueFalse=1)
            for pon in p:
                poc=poc+1
        print("el numero de ponencias rechazadas son: ",poc)
        results=[]
        doctor_json={}
        doctor_json['revistas']=r
        doctor_json['articulos']=ar
        doctor_json['ponencias']=poc
        #print("el numero de revistas rechazadas son",doctor_json['estado'])
        results.append(doctor_json)
        data_json=json.dumps(results)
    else:data_json='fail'
    mimetype="application/json"
    return HttpResponse(data_json,mimetype)

#Clase agregado por: Quinchimbla; Azogue; Caiza; Guilcazo; Potosi; Jaya
#Descargar un reporte de excel por Investigador  fecha: 23/07/2019

class Reporte_excel(TemplateView):
    def get(self, request, *args, **kwargs):
        usuaruia = self.request.user.id
        articulos = articulos_cientificos.objects.filter(user_id=usuaruia)
        libros = libro.objects.filter()
        wb = Workbook()
        ws = wb.active
        ws.title = "ARTÍCULOS"
        ws1 = wb.create_sheet("LIBROS")
        ws3 = wb.create_sheet("CAPÍTULOS")
        ws2 = wb.create_sheet("PONENCIAS")
        TITULO = Font(
            name='Times New Roman',
            size=9,
            bold=True,
        )
        SUBTITULO = Font(
            name='Times New Roman',
            size=8,
            bold=True,
        )
        ws['A1'] = 'REPORTE DE ARTÍCULOS'
        ws.merge_cells('A1:D1')
        ws['A3'] = 'CÓDIGO ASIGNADO POR LA DRECCIÓN DE INVESTIGACIÓN'
        ws['B3'] = '   AÑO'
        ws['C3'] = '   INDEXACIÓN'
        ws['D3'] = 'CÓDIGO DOI / ASIGNADO POR LA DIRECCIÓN DE INVESTIGACIÓN '
        ws['E3'] = '   NOMBRE DE LA PUBLICACIÓN'
        ws['F3'] = 'CÓDIGO ISSN'
        ws['G3'] = 'NOMBRE DE LA REVISTA'
        ws['H3'] = 'NÚMERO REVISTA: '
        ws['I3'] = 'SJR'
        ws['J3'] = 'FECHA_PUBLICACIÓN'
        ws['K3'] = 'CAMPO AMPLIO'
        ws['L3'] = 'CAMPO ESPECÍFICO'
        ws['M3'] = 'CAMPO_DETALLADO'
        ws['N3'] = 'FILIACIÓN'
        ws['O3'] = 'ESTADO'
        ws['P3'] = 'LINK_PUBLICACIÓN'
        ws['Q3'] = 'LINK_REVISTA'
        ws['R3'] = 'FILIACIÓN'
        ws['S3'] = 'N° CÉDULA'
        ws['T3'] = 'AUTOR'
        ws['U3'] = 'N° CÉDULA'
        ws['V3'] = 'AUTOR'
        ws['W3'] = 'N° CÉDULA'
        ws['X3'] = 'AUTOR'
        ws['Y3'] = 'N° CÉDULA'
        ws['Z3'] = 'AUTOR'
        ws['AA3'] = 'N° CÉDULA'
        ws['AB3'] = 'AUTOR'
        ws['AC3'] = 'N° CÉDULA'
        ws['AD3'] = 'AUTOR'
        ws['AE3'] = 'N° CÉDULA'
        ws['AF3'] = 'AUTOR'
        ws['AG3'] = 'N° CÉDULA'
        ws['AH3'] = 'AUTOR'
        ws['AI3'] = 'N° CÉDULA'
        ws['AJ3'] = 'AUTOR'
        ws['AK3'] = 'N° CÉDULA'
        ws['AL3'] = 'AUTOR'
        ws['AM3'] = 'N° CÉDULA'
        ws['AN3'] = 'AUTOR'
        ws.column_dimensions['A'].width = float(26)
        ws.column_dimensions['B'].width = float(11)
        ws.column_dimensions['C'].width = float(13)
        ws.column_dimensions['D'].width = float(41)
        ws.column_dimensions['E'].width = float(48)
        ws.column_dimensions['F'].width = float(16)
        ws.column_dimensions['G'].width = float(42)
        ws.column_dimensions['H'].width = float(12)
        ws.column_dimensions['I'].width = float(20)
        ws.column_dimensions['J'].width = float(19)
        ws.column_dimensions['K'].width = float(36)
        ws.column_dimensions['L'].width = float(36)
        ws.column_dimensions['M'].width = float(36)
        ws.column_dimensions['N'].width = float(10)
        ws.column_dimensions['O'].width = float(11)
        ws.column_dimensions['P'].width = float(31)
        ws.column_dimensions['Q'].width = float(31)
        ws.column_dimensions['R'].width = float(11)
        ws.column_dimensions['S'].width = float(11)
        ws.column_dimensions['T'].width = float(31)
        ws.column_dimensions['U'].width = float(11)
        ws.column_dimensions['V'].width = float(31)
        ws.column_dimensions['W'].width = float(11)
        ws.column_dimensions['X'].width = float(31)
        ws.column_dimensions['Y'].width = float(11)
        ws.column_dimensions['Z'].width = float(31)
        ws.column_dimensions['AA'].width = float(11)
        ws.column_dimensions['AB'].width = float(31)
        ws.column_dimensions['AC'].width = float(11)
        ws.column_dimensions['AD'].width = float(31)
        ws.column_dimensions['AE'].width = float(11)
        ws.column_dimensions['AF'].width = float(31)
        ws.column_dimensions['AG'].width = float(11)
        ws.column_dimensions['AH'].width = float(31)
        ws.column_dimensions['AI'].width = float(11)
        ws.column_dimensions['AJ'].width = float(31)
        ws.row_dimensions[3].height = float(21.75)
        ws['A1'].font = TITULO
        alineacion = Alignment(horizontal='center', vertical='justify')

        for rows in ws.iter_cols(min_col=1, max_col=43, min_row=3, max_row=3):
            for j in rows:
                j.font = SUBTITULO
                j.alignment = alineacion

        cont = 5
        autart = autoresArticulos.objects.filter(user_id=usuaruia)
        for art in autart:
            x = articulos_cientificos.objects.filter(pk=art.articulo_id)
        items = articulos_cientificos.objects.filter(user_id=usuaruia)
        items1 = revista.objects.filter(user_id=usuaruia)
        items2 = autoresArticulos.objects.filter(user_id=usuaruia)
        items3 = Investigador.objects.filter(user_id=usuaruia)
        for art in autart:
            x = articulos_cientificos.objects.filter(pk=art.articulo_id)
            for obj in x:
                autores = autoresArticulos.objects.filter(articulo_id=obj.id)
                vecCed=[]
                vecAut = []
                for aux in autores:
                    infInves=Investigador.objects.filter(user_id=aux.user.id)
                    for buscarid in infInves:
                        vecAut.append(buscarid.user.first_name +' '+ buscarid.user.last_name)
                        vecCed.append(str(buscarid.cedula))
                        ws.cell(row=cont, column=19).value = "".join(vecCed[0])
                        ws.cell(row=cont, column=20).value = "".join(vecAut[0])
                        ws.cell(row=cont, column=21).value = "".join(vecCed[1:2])
                        ws.cell(row=cont, column=22).value = "".join(vecAut[1:2])
                        ws.cell(row=cont, column=23).value = "".join(vecCed[2:3])
                        ws.cell(row=cont, column=24).value = "".join(vecAut[2:3])
                        ws.cell(row=cont, column=25).value = "".join(vecCed[3:4])
                        ws.cell(row=cont, column=26).value = "".join(vecAut[3:4])
                        ws.cell(row=cont, column=27).value = "".join(vecCed[4:5])
                        ws.cell(row=cont, column=28).value = "".join(vecAut[4:5])
                        ws.cell(row=cont, column=29).value = "".join(vecCed[5:6])
                        ws.cell(row=cont, column=30).value = "".join(vecAut[5:6])
                        ws.cell(row=cont, column=31).value = "".join(vecCed[6:7])
                        ws.cell(row=cont, column=32).value = "".join(vecAut[6:7])
                        ws.cell(row=cont, column=33).value = "".join(vecCed[7:8])
                        ws.cell(row=cont, column=34).value = "".join(vecAut[7:8])
                        ws.cell(row=cont, column=35).value = "".join(vecCed[8:9])
                        ws.cell(row=cont, column=36).value = "".join(vecAut[8:9])
                        ws.cell(row=cont, column=37).value = "".join(vecCed[9:10])
                        ws.cell(row=cont, column=38).value = "".join(vecAut[9:10])
                        ws.cell(row=cont, column=39).value = "".join(vecCed[10:11])
                        ws.cell(row=cont, column=40).value = "".join(vecAut[10:11])
                        ws.cell(row=cont, column=41).value = "".join(vecCed[11:12])
                        ws.cell(row=cont, column=42).value = "".join(vecAut[11:12])
                for obj2 in items3:
                    ws.cell(row=cont, column=1).value = obj.id
                    ws.cell(row=cont, column=2).value = obj.fechaPublicacion
                    ws.cell(row=cont, column=3).value = 'SN'
                    ws.cell(row=cont, column=4).value = obj.doi
                    ws.cell(row=cont, column=5).value = obj.titulo
                    ws.cell(row=cont, column=6).value = obj.iSSN
                    ws.cell(row=cont, column=7).value = obj.revista.Nombre
                    ws.cell(row=cont, column=8).value = obj.revista_id
                    ws.cell(row=cont, column=9).value = obj.revista.Factor_Impacto
                    ws.cell(row=cont, column=10).value = obj.uploaded_at
                    ws.cell(row=cont, column=11).value = obj.amplio.Nombre
                    ws.cell(row=cont, column=12).value = obj.especifico.Nombre
                    ws.cell(row=cont, column=13).value = obj.detallado.Nombre
                    ws.cell(row=cont, column=14).value = obj.filialUtc
                    ws.cell(row=cont, column=15).value = obj.estado
                    ws.cell(row=cont, column=16).value = obj.url
                    ws.cell(row=cont, column=17).value = obj.revista.Url
                    ws.cell(row=cont, column=18).value = obj.filialUtc
                    cont += 1

        ws1['A1'] = 'REPORTE DE LIBROS'
        ws1.merge_cells('A1:D1')
        ws1['A3'] = 'FECHA DE PUBLICACIÓN'
        ws1['B3'] = 'CODIFICACIÓN ANUAL'
        ws1['C3'] = 'CÓDIGO ASIGNADO POR LA DIRECCIÓN DE INVESTIGACIÓN'
        ws1['D3'] = 'NOMBRE DE LA OBRA '
        ws1['E3'] = 'CÓDIGO ISBN IMPRESO'
        ws1['F3'] = 'CÓDIGO ISBN DIGITAL'
        ws1['G3'] = 'REVISADA POR PARES'
        ws1['H3'] = 'FILIACIÓN '
        ws1['I3'] = 'CAMPO AMPLIO'
        ws1['J3'] = 'CAMPO ESPECÍFICO'
        ws1['K3'] = 'CAMPO DETALLADO'
        ws1['L3'] = 'N° CÉDULA'
        ws1['M3'] = 'AUTOR'
        ws1['N3'] = 'N° CÉDULA'
        ws1['O3'] = 'AUTOR'
        ws1['P3'] = 'N° CÉDULA'
        ws1['Q3'] = 'AUTOR'
        ws1['R3'] = 'N° CÉDULA'
        ws1['S3'] = 'AUTOR'
        ws1['T3'] = 'N° CÉDULA'
        ws1['U3'] = 'AUTOR'
        ws1['V3'] = 'N° CÉDULA'
        ws1['W3'] = 'AUTOR'
        ws1['X3'] = 'N° CÉDULA'
        ws1['Y3'] = 'AUTOR'
        ws1['Z3'] = 'N° CÉDULA'
        ws1['AA3'] = 'AUTOR'
        ws1['AB3'] = 'N° CÉDULA'
        ws1['AC3'] = 'AUTOR'
        ws1['AD3'] = 'N° CÉDULA'
        ws1['AE3'] = 'AUTOR'
        ws1['AF3'] = 'N° CÉDULA'
        ws1['AG3'] = 'AUTOR'
        ws1.column_dimensions['A'].width = float(20)
        ws1.column_dimensions['B'].width = float(15)
        ws1.column_dimensions['C'].width = float(26)
        ws1.column_dimensions['D'].width = float(47)
        ws1.column_dimensions['E'].width = float(18)
        ws1.column_dimensions['F'].width = float(14)
        ws1.column_dimensions['G'].width = float(40)
        ws1.column_dimensions['H'].width = float(10)
        ws1.column_dimensions['I'].width = float(36)
        ws1.column_dimensions['J'].width = float(36)
        ws1.column_dimensions['K'].width = float(36)
        ws1.column_dimensions['L'].width = float(11)
        ws1.column_dimensions['M'].width = float(31)
        ws1.column_dimensions['N'].width = float(11)
        ws1.column_dimensions['O'].width = float(31)
        ws1.column_dimensions['P'].width = float(11)
        ws1.column_dimensions['Q'].width = float(31)
        ws1.column_dimensions['R'].width = float(11)
        ws1.column_dimensions['S'].width = float(31)
        ws1.column_dimensions['T'].width = float(11)
        ws1.column_dimensions['U'].width = float(31)
        ws1.column_dimensions['V'].width = float(11)
        ws1.column_dimensions['W'].width = float(31)
        ws1.column_dimensions['X'].width = float(11)
        ws1.column_dimensions['Y'].width = float(31)
        ws1.column_dimensions['Z'].width = float(11)
        ws1.column_dimensions['AA'].width = float(31)
        ws1.column_dimensions['AB'].width = float(11)
        ws1.column_dimensions['AC'].width = float(31)
        ws1.column_dimensions['AD'].width = float(11)
        ws1.column_dimensions['AE'].width = float(31)
        ws1.column_dimensions['AF'].width = float(11)
        ws1.column_dimensions['AG'].width = float(31)
        ws1.row_dimensions[3].height = float(21.75)
        ws1['A1'].font = TITULO
        for rows1 in ws1.iter_cols(min_col=1, max_col=43, min_row=3, max_row=3):
            for j in rows1:
                j.font = SUBTITULO
                j.alignment = alineacion

        cont1 = 5
        autlib = autoresLibro.objects.filter(user_id=usuaruia)

        items6 = Investigador.objects.filter(user_id=usuaruia)
        for art1 in autlib:
            y = libro.objects.filter(pk=art1.libro_id)
            for obj1 in y:
                bc = autoresLibro.objects.filter(libro_id=obj1.id)
                coaA = []
                vecCed1=[]
                for a in bc:
                    infI=Investigador.objects.filter(user_id=a.user.id)
                    for vrid in infI:
                        coaA.append(vrid.user.first_name +' '+ vrid.user.last_name)
                        vecCed1.append(str(vrid.cedula))
                        ws1.cell(row=cont1, column=12).value = "".join(vecCed1[0])
                        ws1.cell(row=cont1, column=13).value = "".join(coaA[0])
                        ws1.cell(row=cont1, column=14).value = "".join(vecCed1[1:2])
                        ws1.cell(row=cont1, column=15).value = "".join(coaA[1:2])
                        ws1.cell(row=cont1, column=16).value = "".join(vecCed1[2:3])
                        ws1.cell(row=cont1, column=17).value = "".join(coaA[2:3])
                        ws1.cell(row=cont1, column=18).value = "".join(vecCed1[3:4])
                        ws1.cell(row=cont1, column=19).value = "".join(coaA[3:4])
                        ws1.cell(row=cont1, column=20).value = "".join(vecCed1[4:5])
                        ws1.cell(row=cont1, column=21).value = "".join(coaA[4:5])
                        ws1.cell(row=cont1, column=22).value = "".join(vecCed1[5:6])
                        ws1.cell(row=cont1, column=23).value = "".join(coaA[5:6])
                        ws1.cell(row=cont1, column=24).value = "".join(vecCed1[6:7])
                        ws1.cell(row=cont1, column=25).value = "".join(coaA[6:7])
                        ws1.cell(row=cont1, column=26).value = "".join(vecCed1[7:8])
                        ws1.cell(row=cont1, column=27).value = "".join(coaA[7:8])
                        ws1.cell(row=cont1, column=28).value = "".join(vecCed1[8:9])
                        ws1.cell(row=cont1, column=29).value = "".join(coaA[8:9])
                        ws1.cell(row=cont1, column=30).value = "".join(vecCed1[9:10])
                        ws1.cell(row=cont1, column=31).value = "".join(coaA[9:10])
                        ws1.cell(row=cont1, column=32).value = "".join(vecCed1[10:11])
                        ws1.cell(row=cont1, column=33).value = "".join(coaA[10:11])
                        ws1.cell(row=cont1, column=34).value = "".join(vecCed1[11:12])
                        ws1.cell(row=cont1, column=35).value = "".join(coaA[11:12])
                for obj5 in items6:
                    ws1.cell(row=cont1, column=1).value = obj1.uploaded_at
                    ws1.cell(row=cont1, column=2).value = obj1.Anio
                    ws1.cell(row=cont1, column=3).value = obj1.id
                    ws1.cell(row=cont1, column=4).value = obj1.Titulo
                    ws1.cell(row=cont1, column=5).value = obj1.ISBN
                    ws1.cell(row=cont1, column=6).value = obj1.Doi
                    ws1.cell(row=cont1, column=7).value = obj1.tipo
                    ws1.cell(row=cont1, column=8).value = obj1.filialUtc
                    ws1.cell(row=cont1, column=9).value = obj1.detallado.especifico.amplio.Nombre
                    ws1.cell(row=cont1, column=10).value = obj1.detallado.especifico.Nombre
                    ws1.cell(row=cont1, column=11).value = obj1.detallado.Nombre
                    cont1 += 1

        ws2['A1'] = 'REPORTE DE PONENCIAS'
        ws2.merge_cells('A1:D1')
        ws2['A3'] = 'AÑO'
        ws2['B3'] = 'NOMBRE DE LA PONENCIA'
        ws2['C3'] = 'INSTITUCIÓN'
        ws2['D3'] = 'TITULO PONENCIA'
        ws2['E3'] = 'CIUDAD'
        ws2['F3'] = 'PAÍS'
        ws2['G3'] = 'CÓDIGO ISBN'
        ws2['H3'] = 'FINANCIAMIENTO'
        ws2['I3'] = 'TIPO DE FUENTE'
        ws2['J3'] = 'FECHA PUBLICACIÓN'
        ws2['K3'] = 'FILIACIÓN'
        ws2['L3'] = 'URL'
        ws2['M3'] = 'NOMBRE DEL ARTÍCULO'
        ws2['N3'] = 'N° CÉDULA'
        ws2['O3'] = 'AUTOR'
        ws2['P3'] = 'N° CÉDULA'
        ws2['Q3'] = 'AUTOR'
        ws2['R3'] = 'N° CÉDULA'
        ws2['S3'] = 'AUTOR'
        ws2['T3'] = 'N° CÉDULA'
        ws2['U3'] = 'AUTOR'
        ws2['V3'] = 'N° CÉDULA'
        ws2['W3'] = 'AUTOR'
        ws2['X3'] = 'N° CÉDULA'
        ws2['Y3'] = 'AUTOR'
        ws2['Z3'] = 'N° CÉDULA'
        ws2['AA3'] = 'AUTOR'
        ws2['AB3'] = 'N° CÉDULA'
        ws2['AC3'] = 'AUTOR'
        ws2['AD3'] = 'N° CÉDULA'
        ws2['AE3'] = 'AUTOR'
        ws2['AF3'] = 'N° CÉDULA'
        ws2['AG3'] = 'AUTOR'
        ws2['AH3'] = 'N° CÉDULA'
        ws2['AI3'] = 'AUTOR'
        ws2.column_dimensions['A'].width = float(27)
        ws2.column_dimensions['B'].width = float(65)
        ws2.column_dimensions['C'].width = float(17)
        ws2.column_dimensions['D'].width = float(47)
        ws2.column_dimensions['E'].width = float(11)
        ws2.column_dimensions['F'].width = float(12)
        ws2.column_dimensions['G'].width = float(19)
        ws2.column_dimensions['H'].width = float(15.50)
        ws2.column_dimensions['I'].width = float(33)
        ws2.column_dimensions['J'].width = float(19)
        ws2.column_dimensions['K'].width = float(12)
        ws2.column_dimensions['L'].width = float(48)
        ws2.column_dimensions['M'].width = float(35)
        ws2.column_dimensions['N'].width = float(11)
        ws2.column_dimensions['O'].width = float(31)
        ws2.column_dimensions['P'].width = float(11)
        ws2.column_dimensions['Q'].width = float(31)
        ws2.column_dimensions['R'].width = float(11)
        ws2.column_dimensions['S'].width = float(31)
        ws2.column_dimensions['T'].width = float(11)
        ws2.column_dimensions['U'].width = float(31)
        ws2.column_dimensions['V'].width = float(11)
        ws2.column_dimensions['W'].width = float(31)
        ws2.column_dimensions['X'].width = float(11)
        ws2.column_dimensions['Y'].width = float(31)
        ws2.column_dimensions['Z'].width = float(11)
        ws2.column_dimensions['AA'].width = float(31)
        ws2.column_dimensions['AB'].width = float(11)
        ws2.column_dimensions['AC'].width = float(31)
        ws2.column_dimensions['AD'].width = float(11)
        ws2.column_dimensions['AE'].width = float(31)
        ws2.column_dimensions['AF'].width = float(11)
        ws2.column_dimensions['AG'].width = float(31)
        ws3.column_dimensions['AH'].width = float(11)
        ws3.column_dimensions['AI'].width = float(31)
        ws2.row_dimensions[3].height = float(21.75)
        ws2['A1'].font = TITULO
        for rows2 in ws2.iter_cols(min_col=1, max_col=43, min_row=3, max_row=3):
            for j2 in rows2:
                j2.font = SUBTITULO
                j2.alignment = alineacion

        cont3 = 5
        contAux=5
        autpon = autoresPonencia.objects.filter(user_id=usuaruia)
        for art3 in autpon:
            y3 = ponencia.objects.filter(pk=art3.ponencia_id)
            for obj10 in y3:
                auxTipo=obj10.tipo
                paisss=pais.objects.filter(pk=obj10.pais_id)
                ciudaddd=ciudad.objects.filter(pk=obj10.ciudad_id)
                capi4 = autoresPonencia.objects.filter(ponencia_id=obj10.id)
                vecCed4=[]
                vecAut4= []
                for obj100 in paisss:
                    autart_nombre = autoresArticulos.objects.filter(articulo_id=obj100.id)
                    ws2.cell(row=cont3, column=5).value = obj100.Nombre
                for objaux in ciudaddd:
                    ws2.cell(row=cont3, column=6).value = objaux.Nombre
                for ac4 in capi4:
                    infInves4=Investigador.objects.filter(user_id=ac4.user.id)
                    autart_nombre = autoresArticulos.objects.filter(user_id=ac4.user.id)
                    for buscarid4 in infInves4:
                        vecAut4.append(buscarid4.user.first_name +' '+ buscarid4.user.last_name)
                        vecCed4.append(str(buscarid4.cedula))
                        ws2.cell(row=cont3, column=14).value = "".join(vecCed4[0])
                        ws2.cell(row=cont3, column=15).value = "".join(vecAut4[0])
                        ws2.cell(row=cont3, column=16).value = "".join(vecCed4[1:2])
                        ws2.cell(row=cont3, column=17).value = "".join(vecAut4[1:2])
                        ws2.cell(row=cont3, column=18).value = "".join(vecCed4[2:3])
                        ws2.cell(row=cont3, column=19).value = "".join(vecAut4[2:3])
                        ws2.cell(row=cont3, column=20).value = "".join(vecCed4[3:4])
                        ws2.cell(row=cont3, column=21).value = "".join(vecAut4[3:4])
                        ws2.cell(row=cont3, column=22).value = "".join(vecCed4[4:5])
                        ws2.cell(row=cont3, column=23).value = "".join(vecAut4[4:5])
                        ws2.cell(row=cont3, column=24).value = "".join(vecCed4[5:6])
                        ws2.cell(row=cont3, column=25).value = "".join(vecAut4[5:6])
                        ws2.cell(row=cont3, column=26).value = "".join(vecCed4[6:7])
                        ws2.cell(row=cont3, column=27).value = "".join(vecAut4[6:7])
                        ws2.cell(row=cont3, column=28).value = "".join(vecCed4[7:8])
                        ws2.cell(row=cont3, column=29).value = "".join(vecAut4[7:8])
                        ws2.cell(row=cont3, column=30).value = "".join(vecCed4[8:9])
                        ws2.cell(row=cont3, column=31).value = "".join(vecAut4[8:9])
                        ws2.cell(row=cont3, column=32).value = "".join(vecCed4[9:10])
                        ws2.cell(row=cont3, column=33).value = "".join(vecAut4[9:10])
                        ws2.cell(row=cont3, column=34).value = "".join(vecCed4[10:11])
                        ws2.cell(row=cont3, column=35).value = "".join(vecAut4[10:11])
                        ws2.cell(row=cont3, column=36).value = "".join(vecCed4[11:12])
                        ws2.cell(row=cont3, column=37).value = "".join(vecAut4[11:12])
                ws2.cell(row=cont3, column=1).value = obj10.id
                ws2.cell(row=cont3, column=2).value = obj10.nombrePonencia
                ws2.cell(row=cont3, column=3).value = obj10.lugarPonencia
                ws2.cell(row=cont3, column=4).value = obj10.tituloPonencia
                ws2.cell(row=cont3, column=7).value = obj10.isbn
                ws2.cell(row=cont3, column=8).value = obj10.financiamiento
                if auxTipo=="1":
                    ws2.cell(row=cont3, column=9).value ='Ponencia sin publicación'
                elif auxTipo=="2":
                    ws2.cell(row=cont3, column=9).value ='Libro de resumenes'
                elif auxTipo=="3":
                    ws2.cell(row=cont3, column=9).value ='Libro de Memorias'
                elif auxTipo=="4":
                    ws2.cell(row=cont3, column=9).value ='Libro de memorias (Con ISBN, publicado en una revista indexada a una base de datos y revisada por pares externo)'
                ws2.cell(row=cont3, column=10).value = obj10.fechaPonencia
                ws2.cell(row=cont3, column=11).value = obj10.filialUtc
                ws2.cell(row=cont3, column=12).value = obj10.urlPonencia
                ws2.cell(row=cont3, column=13).value = ''
                cont3 += 1

        #Activar cuando la base de datos de Capitulos este habilitada en ECUCIENCIA
        """ws3['A1'] = 'REPORTE DE CAPÍTULOS'
        ws3.merge_cells('A1:D1')
        ws3['A3'] = 'FECHA PUBLICACIÓN'
        ws3['B3'] = 'CODIFICACIÓN ANUAL'
        ws3['C3'] = 'CÓDIGO ASIGNADO POR LA DRECCIÓN DE INVESTIGACIÓN'
        ws3['D3'] = 'NUMERO DEL CAPÍTULO'
        ws3['E3'] = 'TÍTULO CAPÍTULO'
        ws3['F3'] = 'TÍTULO LIBRO'
        ws3['G3'] = 'CÓDIGO ISBN IMPRESO'
        ws3['H3'] = 'CÓDIGO ISBN DIGITAL'
        ws3['I3'] = 'EDITOR O COMPILADOR'
        ws3['J3'] = 'FECHA PUBLICACIÓN'
        ws3['K3'] = 'PAGINA(S)'
        ws3['L3'] = 'CAMPO AMPLIO'
        ws3['M3'] = 'CAMPO ESPECÍFICO'
        ws3['N3'] = 'CAMPO DETALLADO'
        ws3['O3'] = 'FILIACIÓN'
        ws3['P3'] = 'ESTADO'
        ws3['Q3'] = 'N° CÉDULA'
        ws3['R3'] = 'AUTOR'
        ws3['S3'] = 'N° CÉDULA'
        ws3['T3'] = 'AUTOR'
        ws3['U3'] = 'N° CÉDULA'
        ws3['V3'] = 'AUTOR'
        ws3['W3'] = 'N° CÉDULA'
        ws3['X3'] = 'AUTOR'
        ws3['Y3'] = 'N° CÉDULA'
        ws3['Z3'] = 'AUTOR'
        ws3['AA3'] = 'N° CÉDULA'
        ws3['AB3'] = 'AUTOR'
        ws3['AC3'] = 'N° CÉDULA'
        ws3['AD3'] = 'AUTOR'
        ws3['AE3'] = 'N° CÉDULA'
        ws3['AF3'] = 'AUTOR'
        ws3['AG3'] = 'N° CÉDULA'
        ws3['AH3'] = 'AUTOR'
        ws3['AI3'] = 'N° CÉDULA'
        ws3['AJ3'] = 'AUTOR'
        ws3['AK3'] = 'N° CÉDULA'
        ws3['AL3'] = 'AUTOR'

        ws3.column_dimensions['A'].width = float(12)
        ws3.column_dimensions['B'].width = float(15)
        ws3.column_dimensions['C'].width = float(26)
        ws3.column_dimensions['D'].width = float(12)
        ws3.column_dimensions['E'].width = float(40)
        ws3.column_dimensions['F'].width = float(40)
        ws3.column_dimensions['G'].width = float(13)
        ws3.column_dimensions['H'].width = float(13)
        ws3.column_dimensions['I'].width = float(15)
        ws3.column_dimensions['J'].width = float(14)
        ws3.column_dimensions['K'].width = float(10)
        ws3.column_dimensions['L'].width = float(48)
        ws3.column_dimensions['M'].width = float(48)
        ws3.column_dimensions['N'].width = float(40)
        ws3.column_dimensions['O'].width = float(10)
        ws3.column_dimensions['P'].width = float(18)
        ws3.column_dimensions['Q'].width = float(11)
        ws3.column_dimensions['R'].width = float(31)
        ws3.column_dimensions['S'].width = float(11)
        ws3.column_dimensions['T'].width = float(31)
        ws3.column_dimensions['U'].width = float(11)
        ws3.column_dimensions['V'].width = float(31)
        ws3.column_dimensions['W'].width = float(11)
        ws3.column_dimensions['X'].width = float(31)
        ws3.column_dimensions['Y'].width = float(11)
        ws3.column_dimensions['Z'].width = float(31)
        ws3.column_dimensions['AA'].width = float(11)
        ws3.column_dimensions['AB'].width = float(31)
        ws3.column_dimensions['AC'].width = float(11)
        ws3.column_dimensions['AD'].width = float(31)
        ws3.column_dimensions['AE'].width = float(11)
        ws3.column_dimensions['AF'].width = float(31)
        ws3.column_dimensions['AG'].width = float(11)
        ws3.column_dimensions['AH'].width = float(31)
        ws3.column_dimensions['AI'].width = float(11)
        ws3.column_dimensions['AJ'].width = float(31)
        ws3.column_dimensions['AK'].width = float(11)
        ws3.column_dimensions['AL'].width = float(31)

        ws3.row_dimensions[3].height = float(21.75)
        ws3['A1'].font = TITULO

        for rows3 in ws3.iter_cols(min_col=1, max_col=43, min_row=3, max_row=3):
            for j3 in rows3:
                j3.font = SUBTITULO
                j3.alignment = alineacion

        cont4 = 5
        autcap = autoresCapitulos.objects.filter(user_id=usuaruia)
        items30 = Investigador.objects.filter(user_id=usuaruia)

        for obj3 in autcap:
            cap3 = Capitulos.objects.filter(pk=obj3.Capitulos_id)
            for aux3 in cap3:
                capi3 = autoresCapitulos.objects.filter(Capitulos_id=aux3.id)
                vecCed3=[]
                vecAut3= []

                for ac3 in capi3:
                    ws3.cell(row=cont4, column=4).value = ac3.capituloNumero
                    ws3.cell(row=cont4, column=5).value = ac3.capituloTitulo
                    infInves3=Investigador.objects.filter(user_id=ac3.user.id)
                    for buscarid3 in infInves3:
                        vecAut3.append(buscarid3.user.first_name +' '+ buscarid3.user.last_name)
                        vecCed3.append(str(buscarid3.cedula))
                        ws3.cell(row=cont4, column=17).value = "".join(vecCed3[0])
                        ws3.cell(row=cont4, column=18).value = "".join(vecAut3[0])
                        ws3.cell(row=cont4, column=19).value = "".join(vecCed3[1:2])
                        ws3.cell(row=cont4, column=20).value = "".join(vecAut3[1:2])
                        ws3.cell(row=cont4, column=21).value = "".join(vecCed3[2:3])
                        ws3.cell(row=cont4, column=22).value = "".join(vecAut3[2:3])
                        ws3.cell(row=cont4, column=23).value = "".join(vecCed3[3:4])
                        ws3.cell(row=cont4, column=24).value = "".join(vecAut3[3:4])
                        ws3.cell(row=cont4, column=25).value = "".join(vecCed3[4:5])
                        ws3.cell(row=cont4, column=26).value = "".join(vecAut3[4:5])
                        ws3.cell(row=cont4, column=27).value = "".join(vecCed3[5:6])
                        ws3.cell(row=cont4, column=28).value = "".join(vecAut3[5:6])
                        ws3.cell(row=cont4, column=29).value = "".join(vecCed3[6:7])
                        ws3.cell(row=cont4, column=30).value = "".join(vecAut3[6:7])
                        ws3.cell(row=cont4, column=31).value = "".join(vecCed3[7:8])
                        ws3.cell(row=cont4, column=32).value = "".join(vecAut3[7:8])
                        ws3.cell(row=cont4, column=33).value = "".join(vecCed3[8:9])
                        ws3.cell(row=cont4, column=34).value = "".join(vecAut3[8:9])
                        ws3.cell(row=cont4, column=35).value = "".join(vecCed3[9:10])
                        ws3.cell(row=cont4, column=36).value = "".join(vecAut3[9:10])
                        ws3.cell(row=cont4, column=37).value = "".join(vecCed3[10:11])
                        ws3.cell(row=cont4, column=38).value = "".join(vecAut3[10:11])
                        ws3.cell(row=cont4, column=39).value = "".join(vecCed3[11:12])
                        ws3.cell(row=cont4, column=40).value = "".join(vecAut3[11:12])
                for obje3 in items30:
                    ws3.cell(row=cont4, column=1).value = aux3.fechaPublicacion
                    ws3.cell(row=cont4, column=2).value = aux3.Anio
                    ws3.cell(row=cont4, column=3).value = aux3.id
                    ws3.cell(row=cont4, column=6).value = aux3.Titulo
                    ws3.cell(row=cont4, column=7).value = aux3.ISBN
                    ws3.cell(row=cont4, column=8).value = aux3.Doi
                    ws3.cell(row=cont4, column=9).value = ''
                    ws3.cell(row=cont4, column=10).value = aux3.fechaPublicacion
                    ws3.cell(row=cont4, column=11).value = ''
                    ws3.cell(row=cont4, column=12).value = aux3.detallado.especifico.amplio.Nombre
                    ws3.cell(row=cont4, column=13).value = aux3.detallado.especifico.Nombre
                    ws3.cell(row=cont4, column=14).value = aux3.detallado.Nombre
                    ws3.cell(row=cont4, column=15).value = aux3.filialUtc
                    ws3.cell(row=cont4, column=16).value = aux3.estado
                    cont4+=1"""

        nombre_archivo = "Reporte.xlsx"
        response = HttpResponse(content_type="aplication/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
