# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import unicodedata
from django.contrib import messages
from django.shortcuts import render, HttpResponseRedirect, get_object_or_404
#from pip._vendor.appdirs import unicode
from django.template.loader import render_to_string
from apps.Articulos_Cientificos.models import articulos_cientificos
from apps.Articulos_Cientificos.form import articuloformdisabled
from apps.Articulos_Cientificos.form import articuloform
from django.core.urlresolvers import reverse_lazy
from django.views.generic import CreateView, ListView, UpdateView, DeleteView,TemplateView
from apps.roles.models import Rol
from apps.Investigador.models import Investigador
from apps.autoresArticulos.models import autoresArticulos
from apps.autoresArticulos.form import autoresForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from apps.palabraClave.models import palabraClave
from apps.Linea_Investigacion.models import linea_investigacion
from apps.Sub_Lin_Investigacion.models import sub_lin_investigacion
from apps.Revista.models import revista
from apps.baseDatos.models import baseDatos
from apps.campoAmplio.models import campoAmplio
from apps.pais.models import pais 
from django.http import JsonResponse, HttpResponse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

""" ------------------------importar para el excel"""
from openpyxl import Workbook
from django.http.response import HttpResponse
# Create your views here.

def index(request):
    objs = palabraClave.objects.all().values('Termino')
    onjsList = list(objs)
    return JsonResponse(onjsList, safe=False)


def deleteArt(request, idArt):
    articulo = articulos_cientificos.objects.get(id = idArt)
    autoresArticulos.objects.filter(articulo = articulo).delete()
    articulo.delete()
    messages.success(request, ('Artículo eliminado.'))
    return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))



def listarArticulo(request):
    return render(request, 'ArticuloCientifico/ListarArticulos.html')

    #Creado por: Pilapaña, Guanotasig, Peralta, Pilatasig, Achote
    #Funcion para enviar correo al crear Nuevo Articulo Cientifico  Fecha: 23/07/2019
def correo(test1, test2):
    us = User.objects.get(id=test1)
    msg = MIMEMultipart()
    message = "El Sr(a). " + us.username + " a registrado una nueva ponencia, con el Titulo:" + test2
    email_content = "El/La Investigador(a): " + us.username + " a registrado un nuevo Artículo Científico, con el Título: " + test2 + """
    <html>

   <head>
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">

   <title>Ecuciencia</title>
   <style type="text/css">
    a {color: #943126;}
    b {color: #ffffff;font-size: 24px; font-family: Chercán}
    body, #header h1, #header h2, p {margin: 0; padding: 0;}
    #main {border: 1px solid #cfcece;}
    img {display: block;}
    #top-message p, #bottom p {color: #3f4042; font-size: 12px; font-family: Arial, Helvetica, sans-serif; }
    #header h1 {color: #f8f9f9 !important; font-family: "Lucida Grande", sans-serif; font-size: 24px; margin-bottom: 0!important; padding-bottom: 0; }
    #header p {color: ##F64C4C !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; font-size: 12px;  }
    h5 {margin: 0 0 0.8em 0;}
    h5 {font-size: 20px; color: ##F64C4C !important; font-family: Arial, Helvetica, sans-serif; }
    p {font-size: 12px; color: ##F64C4C !important; font-family: "Lucida Grande", "Lucida Sans", "Lucida Sans Unicode", sans-serif; line-height: 1.5;}
   </style>
     <style type="text/css">
    .boton_personalizado{
    text-decoration: none;
    padding: 8px;
    font-weight: 600;
    font-size: 12px;
    color: #ffffff;
    background-color: #6388CD;
    border-radius: 6px;
  }
</style>
</head>

<body>


<table id="top-message" cellpadding="20" cellspacing="0" width="600" align="center">
    <tr>

    </tr>
  </table>

    <tr><table id="main" width="100" align="center" cellpadding="0" cellspacing="5" bgcolor="6388CD">

      <td align="center">

        <b>ECUCIENCIA</b>

        <table id="header"  cellpadding="10" cellspacing="0" align="center" bgcolor="F0F3F4">
          <tr>
             <td>
        <table>
          <tr>
            <td width="200" valign="top">
              <img src="http://ecuciencia.utc.edu.ec/static/images/logo4151.png" width="230" height="100">
            </td>
            <td width="15"></td>
            <td width="200" valign="top" align="center">
              <h5>Lista de Nuevos Artículos Científicos </h5>
              <a class="boton_personalizado"  href="http://ecuciencia.utc.edu.ec/articulo/ListarArticulos">INGRESAR </a>
            </td>
          </tr>
        </table>
      </td>
          </tr>
        </table>
      </td>
    </tr>

  </table>
  <table id="button" cellpadding="20" cellspacing="0" width="600" align="center">
    <tr>
      <td align="center">
        <p>Este mensaje fue enviado a  por EcuCiencia</p>
        <p><a href="#">ecu.ciencia@utc.edu.ec</a> </p>
      </td>
    </tr>
  </table><!-- top message -->
</td></tr></table><!-- wrapper -->

</body>
    </html>


    """

    # setup the parameters of the message
    password = "3cuci3ncia.2018"
    msg['From'] = "ecu.ciencia@utc.edu.ec"
    msg['To'] = "alex.cevallos@utc.edu.ec, jessy.lema@utc.edu.ec"  # SECRETAIA DE INVESTIGACION
    msg['Subject'] = "ARTÍCULO: EL/LA INVESTIGADOR(A) " + us.first_name + " " + us.last_name + " REGISTRA UN NUEVO ARTÍCULO CIENTÍFICO"

    # add in the message body
    part1 = MIMEText(message, 'plain')
    part2 = MIMEText(email_content, 'html')

    msg.attach(part1)
    msg.set_payload(part2)

    # create server
    server = smtplib.SMTP('smtp.gmail.com: 587')

    server.starttls()

    # Login Credentials for sending the mail
    server.login(msg['From'], password)

    # send the message via the server.
    server.sendmail(msg['From'], msg['To'].split(","), msg.as_string())

    server.quit()

    print
    "successfully sent email to %s:" % (msg['To'])


def elimina_tildes(cadena):
    s = ''.join((c for c in unicodedata.normalize('NFD', cadena) if unicodedata.category(c) != 'Mn'))
    return s


class articulocreate(CreateView):
    model = articulos_cientificos
    model_second = autoresArticulos
    form_class = articuloform

    template_name = 'ArticuloCientifico/CreateArticulo.html'
    success_url = reverse_lazy('articulosCientificos:ListaArticulo')
    def get_context_data(self, **kwargs):
        usuario = self.request.user.id
        context = super(articulocreate, self).get_context_data(**kwargs)
        perfil = Investigador.objects.get(user_id=self.request.user.id)
        US = []
        for i in User.objects.all():
            US.append(i)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        context['us'] = User.objects.exclude(is_superuser = True, is_staff = True).order_by('last_name')
        context['linea'] = linea_investigacion.objects.all()
        context['subLinea'] = sub_lin_investigacion.objects.all()
        context['base'] = baseDatos.objects.all()
        context['revista'] = revista.objects.all()
        context['Pon']=revista.objects.filter(user_id=usuario, estado=4)
        context['amplio'] = campoAmplio.objects.all()
        context['p'] = pais.objects.all().order_by('Nombre')
        context['perfil'] = perfil
        return context

    def post(self, request, *args, **kwargs):

        self.object = self.get_object
        usuario = self.request.user.id
        form = self.form_class(request.POST, request.FILES)
        
        if form.is_valid():

            userList=[]
            for i in User.objects.all():
                json = {}
                json['name'] =i.last_name + ' ' + i.first_name
                json['id'] = i.id
                userList.append(json)

            p = request.POST.getlist('palabraC')
            palabras = []
            for i in p:
                palabras.append(elimina_tildes(i.lower()))
            palabras = list(set(palabras))

            articulo = form.save(commit=False)
            articulo.save()
            form.save_m2m()

            for i in range(7):
                grado = "grado"+str(i)
                user = ''+str(i)
                if grado in request.POST:
                    usId = 0
                    for i in userList: 
                        if i['name'] == request.POST[user]:
                            usId = i['id'] 
                    if usId != 0:
                        a = request.POST[grado]
                        obj = User.objects.get(pk=int(usId))
                        autor = autoresArticulos.objects.create(gradoAutoria=a, articulo=articulo, user=obj)
                        autor.save()
                        del autor


            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    articulo.palabraClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    articulo.palabraClave.add(palabra)
            
            articulo.user_id = usuario
            titul = articulo.titulo
            articulo.save()

            articulo.AMtitulo = elimina_tildes(articulo.titulo.upper())
            articulo.AMresumen = elimina_tildes(articulo.resumen.upper())
            articulo.save()

            comprobacion = autoresArticulos.objects.filter(articulo_id = articulo.id)

            if comprobacion.count() == 0:
                autor = autoresArticulos.objects.create(gradoAutoria="Primero", articulo=articulo, user_id=usuario)
                autor.save()
                messages.warning(request, ('Se registro el articulo, pero ocurrió un error con los autores revise la información nuevamente'))
                return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))

            correo(usuario,titul) #Llama a la funcion correo al guardar el articulo
            messages.success(request, ('Se registro el articulo correctamente'))
            return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))
        else:
            messages.error(request, ('Ha ocurrido un error, revise la información del formulario nuevamente.'))
            return self.render_to_response(self.get_context_data(form=form))


class articuloUpdate(UpdateView):
    model = articulos_cientificos
    form_class = articuloform
    template_name = 'ArticuloCientifico/UpdateArticulo.html'
    success_url = reverse_lazy('articulosCientificos:ListaArticulo')
    def get_context_data(self, **kwargs):
        context = super(articuloUpdate, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        ArticuloID = self.object.id
        articulo = articulos_cientificos.objects.get(id=ArticuloID)
        basesD = [val.id for val in articulo.baseDatos.all()]
        tipoBase = articulo.baseDatos.all().first()

        users = [i.user_id for i in autoresArticulos.objects.filter(articulo = articulo).order_by('id')]
        print(articulo.university.pais_id)
        if not  autoresArticulos.objects.filter(articulo = articulo, user_id = self.request.user.id):
            context['permisos'] = 'invalido'
        perfil = Investigador.objects.get(user_id=self.request.user.id)
        palabra = [val.id for val in articulo.palabraClave.all()]
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        context['Autores'] = autoresArticulos.objects.filter(articulo=articulo).order_by('id')
        context['users'] = User.objects.filter(is_superuser = False, is_staff = False, id__in = users)
        context['Datos'] = baseDatos.objects.exclude(id__in=basesD)

        context['DB'] = articulo.baseDatos.all()
        context['DB1'] = tipoBase.tipoBaseDatos_id 
        #Cambios realizados por: David Guaman & Fernanda Barragan

        context['DB2'] = [val.id for val in articulo.baseDatos.all()]
        #tipoBase.tipoBaseDatos_id

        context['palabra'] = palabraClave.objects.filter(id__in=palabra)
        context['linea'] = linea_investigacion.objects.all()
        context['subLinea'] = sub_lin_investigacion.objects.all()
        context['base'] = baseDatos.objects.all()
        context['revista'] = revista.objects.all()
        context['Pon']=revista.objects.filter(user_id=usuario, estado=4)
        context['id'] = ArticuloID
        context['perfil'] = perfil
        context['amplio'] = campoAmplio.objects.all()
        context['p'] = pais.objects.all().order_by('Nombre')
        context['paisId'] = articulo.university.pais_id
        return context


    def post(self, request, *args, **kwargs):
        self.object = self.get_object

        usuario = self.request.user.id
        id_art = kwargs['pk']
        art = self.model.objects.get(id=id_art)
        form = self.form_class(request.POST, request.FILES, instance=art)
        
        if form.is_valid():
            userList = []
            for i in User.objects.all():
                json = {}
                json['name'] = i.last_name + ' ' + i.first_name
                json['id'] = i.id
                userList.append(json)
            p = request.POST.getlist('palabras')
            palabras = []
            for i in p:
                palabras.append(elimina_tildes(i.lower()))
            palabras = list(set(palabras))
            articulo = form.save(commit=False)
            articulo.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    articulo.palabraClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    articulo.palabraClave.add(palabra)
            au = autoresArticulos.objects.filter(articulo = art)
            for i in au:
                i.delete()
            
            for i in range(7):
                grado = "grado"+str(i)
                user = ""+str(i)
                if grado in request.POST:
                    usId = 0
                    for i in userList:
                        if i['name'] == request.POST[user]:
                            usId = i['id']
                    if usId != 0:
                        a = request.POST[grado]
                        obj = User.objects.get(pk=int(usId))
                        autor = autoresArticulos.objects.create(gradoAutoria=a, articulo=articulo, user=obj)
                        autor.save()
                        del autor
            
            articulo.statusText = False
            articulo.save()

            if articulo.textMining != None:
                articulo.textMining.statusClasificacion = False
                articulo.textMining.sublineaClasificacion = None
                articulo.textMining.save()


            articulo.AMtitulo = elimina_tildes(articulo.titulo.upper())
            articulo.AMresumen = elimina_tildes(articulo.resumen.upper())
            articulo.save()
            
            comprobacion = autoresArticulos.objects.filter(articulo_id = articulo.id)

            if comprobacion.count() == 0:
                autor = autoresArticulos.objects.create(gradoAutoria="Primero", articulo=articulo, user_id=articulo.user_id)
                autor.save()
                messages.warning(request, ('Se registro el articulo, pero ocurrió un error con los autores revise la información nuevamente'))
                return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))


            messages.success(request, ('Se ha actualizado la información del articulo correctamente'))
            return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))
        else:
            messages.error(request, ('Ah ocurrido un error, revise la informacion del formulario.'))
            return self.render_to_response(self.get_context_data(form=form))

class articuloUpdateDisabled(UpdateView):
    model = articulos_cientificos
    form_class = articuloformdisabled
    template_name = 'ArticuloCientifico/UpdateArticulo.html'
    success_url = reverse_lazy('articulosCientificos:ListaArticulo')
    def get_context_data(self, **kwargs):
        context = super(articuloUpdateDisabled, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        ArticuloID = self.object.id
        articulo = articulos_cientificos.objects.get(id=ArticuloID)
        basesD = [val.id for val in articulo.baseDatos.all()]
        tipoBase = articulo.baseDatos.all().first()

        users = [i.user_id for i in autoresArticulos.objects.filter(articulo = articulo).order_by('id')]
        print(articulo.university.pais_id)
        if not  autoresArticulos.objects.filter(articulo = articulo, user_id = self.request.user.id):
            context['permisos'] = 'invalido'
        perfil = Investigador.objects.get(user_id=self.request.user.id)
        palabra = [val.id for val in articulo.palabraClave.all()]
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        context['Autores'] = autoresArticulos.objects.filter(articulo=articulo).order_by('id')
        context['users'] = User.objects.filter(is_superuser = False, is_staff = False, id__in = users)
        context['Datos'] = baseDatos.objects.exclude(id__in=basesD)

        context['DB'] = articulo.baseDatos.all()
        context['DB1'] = tipoBase.tipoBaseDatos_id

        context['DB2'] = [val.id for val in articulo.baseDatos.all()]
        #tipoBase.tipoBaseDatos_id

        context['palabra'] = palabraClave.objects.filter(id__in=palabra)
        context['linea'] = linea_investigacion.objects.all()
        context['subLinea'] = sub_lin_investigacion.objects.all()
        context['base'] = baseDatos.objects.all()
        context['revista'] = revista.objects.all()
        context['id'] = ArticuloID
        context['perfil'] = perfil
        context['amplio'] = campoAmplio.objects.all()
        context['p'] = pais.objects.all().order_by('Nombre')
        context['paisId'] = articulo.university.pais_id
        return context


    def post(self, request, *args, **kwargs):
        self.object = self.get_object


        usuario = self.request.user.id
        id_art = kwargs['pk']
        art = self.model.objects.get(id=id_art)
        form = self.form_class(request.POST, request.FILES, instance=art)
        
        if form.is_valid():
            userList = []
            for i in User.objects.all():
                json = {}
                json['name'] = i.last_name + ' ' + i.first_name
                json['id'] = i.id
                userList.append(json)
            p = request.POST.getlist('palabras')
            palabras = []
            for i in p:
                palabras.append(elimina_tildes(i.lower()))
            palabras = list(set(palabras))
            articulo = form.save(commit=False)
            articulo.save()
            form.save_m2m()
            for i in palabras:
                if not palabraClave.objects.filter(Termino__iexact=i):
                    us = User.objects.get(pk=usuario)
                    palabraNew = palabraClave.objects.create(Termino=i, user=us, PCMtermino = i.upper())
                    articulo.palabraClave.add(palabraNew)
                    del palabraNew
                else:
                    palabra = palabraClave.objects.get(Termino__iexact=i)
                    articulo.palabraClave.add(palabra)
            au = autoresArticulos.objects.filter(articulo = art)
            for i in au:
                i.delete()
            
            for i in range(7):
                grado = "grado"+str(i)
                user = ""+str(i)
                if grado in request.POST:
                    usId = 0
                    for i in userList:
                        if i['name'] == request.POST[user]:
                            usId = i['id']
                    if usId != 0:
                        a = request.POST[grado]
                        obj = User.objects.get(pk=int(usId))
                        autor = autoresArticulos.objects.create(gradoAutoria=a, articulo=articulo, user=obj)
                        autor.save()
                        del autor
            
            articulo.save()

            articulo.AMtitulo = elimina_tildes(articulo.titulo.upper())
            articulo.tituloSearch = elimina_tildes(articulo.titulo.upper())
            articulo.AMresumen = elimina_tildes(articulo.resumen.upper())
            articulo.statusText = False
            articulo.save()
            
            if articulo.textMining != None:
                articulo.textMining.statusClasificacion = False
                articulo.textMining.sublineaClasificacion = None
                articulo.textMining.save()
            
            comprobacion = autoresArticulos.objects.filter(articulo_id = articulo.id)

            if comprobacion.count() == 0:
                autor = autoresArticulos.objects.create(gradoAutoria="Primero", articulo=articulo, user_id=articulo.user_id)
                autor.save()
                messages.warning(request, ('Se registro el articulo, pero ocurrió un error con los autores revise la información nuevamente'))
                return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))


            messages.success(request, ('Se ha actualizado la información del articulo correctamente'))
            return HttpResponseRedirect(reverse_lazy('articulosCientificos:ListaArticulo'))
        else:
            messages.error(request, ('Ah ocurrido un error, revise la informacion del formulario.'))
            return self.render_to_response(self.get_context_data(form=form))

class articuloDelete(DeleteView):
    model = articulos_cientificos
    template_name = 'ArticuloCientifico/DeleteArticulo.html'
    success_url = reverse_lazy('articulosCientificos:ListaArticulo')
    def get_context_data(self, **kwargs):
        context = super(articuloDelete, self).get_context_data(**kwargs)
        usuario = self.request.user.id
        perfil = Investigador.objects.get(user_id=usuario)
        roles = perfil.roles.all()
        privi = []
        privilegios = []
        privilegio = []
        for r in roles:
            privi.append(r.id)
        for p in privi:
            roles5 = Rol.objects.get(pk=p)
            priv = roles5.privilegios.all()
            for pr in priv:
                privilegios.append(pr.codename)
        for i in privilegios:
            if i not in privilegio:
                privilegio.append(i)
        if not  autoresArticulos.objects.filter(articulo_id = self.object.id, user_id = self.request.user.id):
            context['permisos'] = 'invalido'  
        context['usuario'] = privilegio

        return context


class articuloList(ListView):
    model = articulos_cientificos
    template_name = 'ArticuloCientifico/ListarArticulos.html'
    # paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super(articuloList, self).get_context_data(**kwargs)
        perfil = Investigador.objects.get(user_id=self.request.user.id)
        us = User.objects.get(id = self.request.user.id)
        aut = autoresArticulos.objects.filter(user = us)
        l = []
        for i in aut:
            l.append(i.articulo.id)
        articulo = articulos_cientificos.objects.filter(id__in = l).order_by('titulo')
        context['articulo'] = articulo
        #Cambios realizados por: Rovayo &Santana &Sarco &Toaquiza &Sandoval &Sanchez
        #Para listar documentos rechazados      Fecha:03/07/2019
        context['articuloAdmin'] = articulos_cientificos.objects.all().exclude(editableTrueFalse=1)
        context['artiiculoAdminRec'] = articulos_cientificos.objects.all()
        context['perfil'] = perfil
        return context

def new2(request):
    x = 0 
    y = 0
    for i in articulos_cientificos.objects.all():
        x+=1
        
        if autoresArticulos.objects.filter(articulo_id = i.id):
            print(x)
        else:
            print(f"{x} Sin autor id {i.id}")
            #print("Sin autor")
        #if 0==i.palabraClave.count():
        #    print(i.palabraClave.count())


    #for i in palabraClave.objects.all():
        """
        i.Termino.replace(".", "")
        i.PCMtermino.replace(".", "")
        i.save()
        """

        """
        x=0
        y=""
        for j in i.Termino:
            if x==0 and j == " ":
                j=""
            y+=j
            #print(j)
            x+=1
        print(y)
        i.Termino = y
        i.save()
        x=0
        y=""
        for j in i.PCMtermino:
            if x==0 and j == " ":
                j=""
            y+=j
            #print(j)
            x+=1
        print(y)
        i.PCMtermino=y
        i.save()
        #print(i.Termino, i.PCMtermino)
        """
        #if i.Termino.find(";") != -1 or i.Termino.find(",") != -1 or i.Termino.find(";") != -1:
        #    print(i.Termino)

    return HttpResponseRedirect('/index/')

#from django_pandas.io import read_frame
#from pandas import read_frame

import goslate
import pandas as pd
#%matplotlib inline
import matplotlib.pyplot as plt
def new(request):
    print(pd.__version__)
    keywords = palabraClave.objects.all()
    #print(keywords.count())
    keywordsRefine = list(set(keywords))
    kw=[]
    #for i in keywordsRefine:
    #    print(i.PCMtermino)
    k=""
    articles = articulos_cientificos.objects.all().order_by('id')
    #gs = goslate.Goslate()
    for i in articles:
        i.titulo
        
        #print(gs.detect(i.titulo))
    
    for i in articles:
        for j in i.palabraClave.all():
            k += f"{j.PCMtermino}"
        kw.append(k)
        k=""

    idArt = [i.id for i in articles]
    l = [i.AMtitulo for i in articles]
    l2= [i.AMresumen for i in articles]
    k1=[]
    k2=[]
    k3=[]
    k4=[]
    k5=[]
    k6=[]
    k7=[]
    k8=[]
    k9=[]
    k10=[]
    k11=[]
    k12=[]
    k13=[]
    k14=[]
    aux=[]
    for i in articles:
        for j in i.palabraClave.all():

            aux.append(j.PCMtermino)
            iaux=0
            listAux = []
            if j.Termino.find(";") != -1 or j.Termino.find(",") != -1:
                iaux=0
                listAux.append(0)
                for k in j.Termino:
                    if k == ";" or k == ",":
                        #print(iaux)
                        listAux.append(iaux)
                    iaux+=1
                
                listAux.append(len(j.Termino))
                total = len(listAux) - 1
                #print(iaux)
                print(j.Termino)
                print(listAux, i.id)
                iaux2 = 0

                for jaux in range(total):
                    if jaux == 0:
                        print(j.Termino[listAux[jaux]:listAux[jaux+1]])
                    else:
                        print(j.Termino[listAux[jaux]+2:listAux[jaux+1]])
                    #print(jaux)
            """
                print(" ")
                print("*--------------------------------------*")
                print(j.PCMtermino)
                
                aux3=""
                for k in j.PCMtermino:
                    aux3 += k
                    if k == ";" or k == "," or k == " - ":
                        print("-------------------------------------------------------")
                        aux3=""
                    print(aux3)
            """  


        if len(aux) <= 14:
            aux2 = 14 - len(aux)
            for j in range(aux2):
                aux.append("")

        k1.append(aux[0])
        k2.append(aux[1])
        k3.append(aux[2])
        k4.append(aux[3])
        k5.append(aux[4])
        k6.append(aux[5])
        k7.append(aux[6])
        k8.append(aux[7])
        k9.append(aux[8])
        k10.append(aux[9])
        k11.append(aux[10])
        k12.append(aux[11])
        k13.append(aux[12])
        k14.append(aux[13])
        aux=[]
    
    articulos = pd.DataFrame(
        {
            "id":idArt,
            "AMtitulo":l,
            "AMresumen":l2,
            "k1":k1,
            "k2":k2,
            "k3":k3,
            "k4":k4,
            "k5":k5,
            "k6":k6,
            "k7":k7,
            "k8":k8,
            "k9":k9,
            "k10":k10,
            "k11":k11,
            "k12":k12,
            "k13":k13,
            "k14":k14,
        }
    )
    articulos = articulos.rename(columns={
        "AMtitulo":"Titulo",
        "AMresumen":"Resumen",
    })
    #articulos.head() ver la cabecera del Dataframe
    #print(articulos.dtypes)
    #print(articulos.shape)
    #print(articulos[articulos.duplicated()].shape)
    
    return HttpResponse(articulos.to_html())
    #return HttpResponseRedirect('/index/')

class ReporteExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        articulos=articulos_cientificos.objects.all()
        wb=Workbook()
        ws=wb.active
        ws['B1']='REPORTE DE ARTÍCULOS CIENTIFICOS'
        ws.merge_cells('B1:E1')
        ws['B3']='TÍTULO'
        ws['C3']='FECHA DE PUBLICACIÓN'
        ws['D3']='ESTADO'
        ws['E3']='ISSN'
        ws['F3']='VOLUMEN'
        ws['G3']='URL'
        cont=6
        for  arti in articulos:
            ws.cell(row=cont,column=2).value=arti.titulo
            ws.cell(row=cont,column=3).value=arti.fechaPublicacion
            ws.cell(row=cont,column=4).value=arti.estado
            ws.cell(row=cont,column=5).value=arti.iSSN
            ws.cell(row=cont,column=6).value=arti.volumen
            ws.cell(row=cont,column=7).value=arti.url
            cont+=1

        nombre_archivo=" ReporteArticulos.xlsx"
        response=HttpResponse(content_type="aplication/ms-excel")
        content="attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

